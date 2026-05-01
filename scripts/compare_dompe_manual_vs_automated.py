"""Compare a manual SOP assessment against our automated evaluator output.

Reads:
  - Automated evaluator results: either a local JSON dump (default) or a
    Supabase ``document_evaluation_id``.
  - Manual ratings: either the already-extracted CSV (default) or the source
    Excel workbook's "Manual Verification" sheet.

Writes per run, into ``docs/docs_for_eval/comparison_<UTC-timestamp>/``:
  - ``comparison.md``      Markdown report with a summary block + per-clause table.
  - ``comparison.xlsx``    Two-sheet workbook with conditional formatting on
                           the agreement column and a full-detail sheet.
  - ``summary.json``       Machine-readable run metadata + counts.

Agreement classification (severity ladder PASS=0, FLAGGED=1, FAIL=2):
  EXACT     mapped manual == auto status
  ADJACENT  off by exactly 1 severity step
  SEVERE    off by 2 or more
  NA        either side is NOT_APPLICABLE (excluded from %)

Usage examples:
  python scripts/compare_dompe_manual_vs_automated.py
  python scripts/compare_dompe_manual_vs_automated.py \\
      --evaluation-json docs/dompe-design-control-simplified-evaluation.json
  python scripts/compare_dompe_manual_vs_automated.py \\
      --evaluation-id <uuid> --evaluator-column "Outlook (PK)"
"""

from __future__ import annotations

import argparse
import csv
import json
import os
import re
import sys
from dataclasses import asdict, dataclass, field
from datetime import datetime, timezone
from pathlib import Path
from typing import Optional

REPO_ROOT = Path(__file__).resolve().parent.parent
DEFAULT_EVAL_JSON = REPO_ROOT / "docs/dompe-design-control-simplified-evaluation.json"
DEFAULT_MANUAL_CSV = REPO_ROOT / "docs/design-control-manual-verification.csv"
DEFAULT_MANUAL_XLSX = REPO_ROOT / "docs/docs_for_eval/Design Control SOP Assessment Tool II (1).xlsx"
DEFAULT_OUT_PARENT = REPO_ROOT / "docs/docs_for_eval"

MANUAL_TO_AUTO = {
    "meets requirement": "PASS",
    "ofi": "FLAGGED",
    "deficiency": "FAIL",
    "not reviewed": "NOT_APPLICABLE",
}
SEVERITY = {"PASS": 0, "FLAGGED": 1, "FAIL": 2}
LABELS = {
    "PASS": "Meets requirement",
    "FLAGGED": "OFI",
    "FAIL": "Deficiency",
    "NOT_APPLICABLE": "Not Reviewed",
    "ERROR": "Error",
    "PARTIAL": "Partial",
}


@dataclass
class ManualEntry:
    clause: str
    title: str
    rating_raw: str
    rating_normalized: str
    notes: str


@dataclass
class AutoEntry:
    clause: str
    title: str
    status: str
    confidence: Optional[str]
    rationale: str
    gaps: str
    recommendations: str


@dataclass
class ComparisonRow:
    clause: str
    title: str
    manual_rating: str
    manual_notes: str
    auto_status: str
    auto_status_label: str
    auto_confidence: str
    agreement: str
    auto_rationale: str
    auto_gaps: str
    auto_recommendations: str


@dataclass
class Summary:
    total: int
    exact: int
    adjacent: int
    severe: int
    not_applicable: int
    exact_pct: float
    adjacent_pct: float
    severe_pct: float
    inputs: dict = field(default_factory=dict)
    framework_slug: str = "design-control-condensed"
    generated_at: str = ""


# ---------- manual loading ----------------------------------------------------

def _split_rating(cell: str) -> tuple[str, str]:
    """Pull the leading rating phrase off a manual cell; return (rating, notes)."""
    cleaned = (cell or "").strip()
    if not cleaned:
        return "", ""
    lines = [ln.rstrip() for ln in cleaned.splitlines()]
    rating = lines[0].strip()
    rest = "\n".join(ln for ln in lines[1:]).strip()
    return rating, rest


def _normalize_manual(rating: str) -> str:
    """Map a manual rating phrase to the auto-evaluator enum, or '' if unknown."""
    return MANUAL_TO_AUTO.get(rating.strip().lower(), "")


def load_manual_csv(path: Path, evaluator_column: str) -> dict[str, ManualEntry]:
    with path.open(newline="", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        if evaluator_column not in reader.fieldnames:
            raise SystemExit(
                f"--evaluator-column {evaluator_column!r} not found in {path}; "
                f"available columns: {reader.fieldnames}"
            )
        out: dict[str, ManualEntry] = {}
        for row in reader:
            clause = (row.get("ID/Clause Number") or "").strip()
            if not clause:
                continue
            rating, notes = _split_rating(row.get(evaluator_column, ""))
            out[clause] = ManualEntry(
                clause=clause,
                title=(row.get("ID/Clause Title") or "").strip(),
                rating_raw=rating,
                rating_normalized=_normalize_manual(rating),
                notes=notes,
            )
        return out


def load_manual_xlsx(path: Path, evaluator_column: str) -> dict[str, ManualEntry]:
    from openpyxl import load_workbook
    wb = load_workbook(path, data_only=True)
    sheet = wb["Manual Verification"]
    headers = [(c.value or "").strip() if isinstance(c.value, str) else c.value for c in sheet[1]]
    try:
        clause_idx = headers.index("ID/Clause Number")
        title_idx = headers.index("ID/Clause Title")
        eval_idx = headers.index(evaluator_column)
    except ValueError as e:
        raise SystemExit(f"Expected column missing in {path}: {e}; headers: {headers}")
    out: dict[str, ManualEntry] = {}
    for row in sheet.iter_rows(min_row=2, values_only=True):
        clause = (row[clause_idx] or "").strip() if isinstance(row[clause_idx], str) else None
        if not clause:
            continue
        rating, notes = _split_rating(str(row[eval_idx] or ""))
        out[clause] = ManualEntry(
            clause=clause,
            title=str(row[title_idx] or "").strip(),
            rating_raw=rating,
            rating_normalized=_normalize_manual(rating),
            notes=notes,
        )
    return out


# ---------- automated loading -------------------------------------------------

def _join_list(value) -> str:
    if not value:
        return ""
    if isinstance(value, list):
        return " | ".join(str(v).strip() for v in value if str(v).strip())
    return str(value).strip()


def load_auto_json(path: Path) -> dict[str, AutoEntry]:
    data = json.loads(path.read_text())
    out: dict[str, AutoEntry] = {}
    for r in data.get("requirements_results", []):
        clause = r.get("requirement_clause") or r.get("clause") or ""
        if not clause:
            continue
        out[clause] = AutoEntry(
            clause=clause,
            title=r.get("requirement_title") or r.get("title") or "",
            status=(r.get("status") or "").upper(),
            confidence=r.get("confidence"),
            rationale=str(r.get("rationale") or "").strip(),
            gaps=_join_list(r.get("gaps") or r.get("gaps_identified")),
            recommendations=_join_list(r.get("recommendations")),
        )
    return out


def load_auto_supabase(evaluation_id: str) -> dict[str, AutoEntry]:
    from dotenv import load_dotenv
    load_dotenv(REPO_ROOT / ".env")
    from supabase import create_client
    url = os.environ.get("SUPABASE_URL")
    key = os.environ.get("SUPABASE_SERVICE_KEY") or os.environ.get("SUPABASE_ANON_KEY")
    if not url or not key:
        raise SystemExit("SUPABASE_URL and SUPABASE_SERVICE_KEY (or ANON) must be set in .env")
    client = create_client(url, key)
    rows = (
        client.table("requirement_evaluations")
        .select(
            "status, confidence_level, evaluation_rationale, gaps_identified, "
            "recommendations, iso_requirements(clause, title, display_order)"
        )
        .eq("document_evaluation_id", evaluation_id)
        .execute()
        .data
    )
    out: dict[str, AutoEntry] = {}
    for r in rows:
        ir = r.get("iso_requirements") or {}
        clause = (ir.get("clause") or "").strip()
        if not clause:
            continue
        out[clause] = AutoEntry(
            clause=clause,
            title=(ir.get("title") or "").strip(),
            status=(r.get("status") or "").upper(),
            confidence=r.get("confidence_level"),
            rationale=str(r.get("evaluation_rationale") or "").strip(),
            gaps=_join_list(r.get("gaps_identified")),
            recommendations=_join_list(r.get("recommendations")),
        )
    return out


# ---------- comparison --------------------------------------------------------

def classify(manual_norm: str, auto_status: str) -> str:
    if manual_norm == "NOT_APPLICABLE" or auto_status == "NOT_APPLICABLE":
        return "NA"
    if manual_norm not in SEVERITY or auto_status not in SEVERITY:
        return "NA"
    delta = abs(SEVERITY[manual_norm] - SEVERITY[auto_status])
    if delta == 0:
        return "EXACT"
    if delta == 1:
        return "ADJACENT"
    return "SEVERE"


def _clause_sort_key(clause: str) -> tuple:
    parts = re.findall(r"\d+", clause)
    return tuple(int(p) for p in parts) if parts else (0,)


def build_rows(
    manual: dict[str, ManualEntry], auto: dict[str, AutoEntry]
) -> list[ComparisonRow]:
    clauses = sorted(set(manual) | set(auto), key=_clause_sort_key)
    rows: list[ComparisonRow] = []
    for clause in clauses:
        m = manual.get(clause)
        a = auto.get(clause)
        if not m and not a:
            continue
        manual_rating = m.rating_raw if m else ""
        manual_notes = m.notes if m else ""
        auto_status = a.status if a else ""
        manual_norm = m.rating_normalized if m else ""
        agreement = classify(manual_norm, auto_status) if (m and a) else "NA"
        rows.append(
            ComparisonRow(
                clause=clause,
                title=(a.title if a else (m.title if m else "")),
                manual_rating=manual_rating,
                manual_notes=manual_notes,
                auto_status=auto_status,
                auto_status_label=LABELS.get(auto_status, auto_status),
                auto_confidence=str(a.confidence) if a and a.confidence else "",
                agreement=agreement,
                auto_rationale=a.rationale if a else "",
                auto_gaps=a.gaps if a else "",
                auto_recommendations=a.recommendations if a else "",
            )
        )
    return rows


def summarize(rows: list[ComparisonRow]) -> Summary:
    total = len(rows)
    exact = sum(1 for r in rows if r.agreement == "EXACT")
    adjacent = sum(1 for r in rows if r.agreement == "ADJACENT")
    severe = sum(1 for r in rows if r.agreement == "SEVERE")
    na = sum(1 for r in rows if r.agreement == "NA")
    denom = max(total - na, 1)
    return Summary(
        total=total,
        exact=exact,
        adjacent=adjacent,
        severe=severe,
        not_applicable=na,
        exact_pct=round(100 * exact / denom, 1),
        adjacent_pct=round(100 * adjacent / denom, 1),
        severe_pct=round(100 * severe / denom, 1),
    )


# ---------- output ------------------------------------------------------------

def _truncate(text: str, n: int = 300) -> str:
    text = (text or "").replace("\n", " ").strip()
    return text if len(text) <= n else text[: n - 1].rstrip() + "…"


def write_markdown(rows: list[ComparisonRow], summary: Summary, out: Path) -> None:
    lines: list[str] = []
    lines.append("# Manual vs Automated Evaluation Comparison")
    lines.append("")
    lines.append(f"Generated: {summary.generated_at}")
    lines.append(f"Framework: `{summary.framework_slug}`")
    lines.append(f"Inputs: {summary.inputs}")
    lines.append("")
    lines.append("## Summary")
    lines.append("")
    denom_note = (
        f" (denominator: {summary.total - summary.not_applicable}; NA: {summary.not_applicable})"
    )
    lines.append(f"- Total clauses: **{summary.total}**{denom_note}")
    lines.append(f"- EXACT:    **{summary.exact}** ({summary.exact_pct}%)")
    lines.append(f"- ADJACENT: **{summary.adjacent}** ({summary.adjacent_pct}%)")
    lines.append(f"- SEVERE:   **{summary.severe}** ({summary.severe_pct}%)")
    lines.append("")
    lines.append("## Per-clause comparison")
    lines.append("")
    headers = [
        "Clause", "Title", "Manual Rating", "Manual Notes",
        "Auto Status", "Auto Confidence", "Agreement",
        "Auto Rationale (truncated)", "Auto Gaps",
    ]
    lines.append("| " + " | ".join(headers) + " |")
    lines.append("| " + " | ".join("---" for _ in headers) + " |")
    for r in rows:
        cells = [
            r.clause,
            r.title,
            r.manual_rating or "—",
            _truncate(r.manual_notes, 200) or "—",
            r.auto_status_label or "—",
            r.auto_confidence or "—",
            r.agreement,
            _truncate(r.auto_rationale, 300) or "—",
            _truncate(r.auto_gaps, 250) or "—",
        ]
        lines.append("| " + " | ".join(_md_cell(c) for c in cells) + " |")
    out.write_text("\n".join(lines) + "\n")


def _md_cell(value: str) -> str:
    """Escape pipes/newlines so a markdown table row stays on one line."""
    return str(value).replace("|", "\\|").replace("\n", " ")


def write_xlsx(rows: list[ComparisonRow], summary: Summary, out: Path) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.formatting.rule import CellIsRule

    wb = Workbook()

    # Sheet 1: Comparison
    ws = wb.active
    ws.title = "Comparison"
    headers = [
        "Clause", "Title", "Manual Rating", "Manual Notes",
        "Auto Status", "Auto Confidence", "Agreement",
        "Auto Rationale (truncated)", "Auto Gaps",
    ]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(vertical="top", wrap_text=True)

    for r in rows:
        ws.append([
            r.clause, r.title, r.manual_rating, r.manual_notes,
            r.auto_status_label, r.auto_confidence, r.agreement,
            _truncate(r.auto_rationale, 300), _truncate(r.auto_gaps, 250),
        ])

    widths = [12, 36, 18, 36, 18, 14, 12, 60, 40]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[chr(64 + i)].width = w
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    last_row = ws.max_row
    if last_row >= 2:
        agreement_range = f"G2:G{last_row}"
        ws.conditional_formatting.add(
            agreement_range,
            CellIsRule(operator="equal", formula=['"EXACT"'],
                       fill=PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")),
        )
        ws.conditional_formatting.add(
            agreement_range,
            CellIsRule(operator="equal", formula=['"ADJACENT"'],
                       fill=PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")),
        )
        ws.conditional_formatting.add(
            agreement_range,
            CellIsRule(operator="equal", formula=['"SEVERE"'],
                       fill=PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")),
        )
        ws.conditional_formatting.add(
            agreement_range,
            CellIsRule(operator="equal", formula=['"NA"'],
                       fill=PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")),
        )

    # Sheet 2: Full Detail
    ws2 = wb.create_sheet("Full Detail")
    ws2.append([
        "Clause", "Title", "Manual Rating", "Manual Notes",
        "Auto Status", "Auto Confidence", "Agreement",
        "Auto Rationale", "Auto Gaps", "Auto Recommendations",
    ])
    for cell in ws2[1]:
        cell.font = Font(bold=True)
    for r in rows:
        ws2.append([
            r.clause, r.title, r.manual_rating, r.manual_notes,
            r.auto_status_label, r.auto_confidence, r.agreement,
            r.auto_rationale, r.auto_gaps, r.auto_recommendations,
        ])
    for i, w in enumerate([12, 36, 18, 50, 18, 14, 12, 80, 60, 60], start=1):
        ws2.column_dimensions[chr(64 + i)].width = w
    for row in ws2.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    # Sheet 3: Summary
    ws3 = wb.create_sheet("Summary")
    ws3.append(["Metric", "Value"])
    ws3["A1"].font = Font(bold=True)
    ws3["B1"].font = Font(bold=True)
    for k, v in [
        ("Generated", summary.generated_at),
        ("Framework", summary.framework_slug),
        ("Total clauses", summary.total),
        ("EXACT", summary.exact),
        ("ADJACENT", summary.adjacent),
        ("SEVERE", summary.severe),
        ("NA (excluded)", summary.not_applicable),
        ("EXACT %", summary.exact_pct),
        ("ADJACENT %", summary.adjacent_pct),
        ("SEVERE %", summary.severe_pct),
    ]:
        ws3.append([k, v])
    ws3.column_dimensions["A"].width = 22
    ws3.column_dimensions["B"].width = 50

    wb.save(out)


def write_summary_json(summary: Summary, out: Path) -> None:
    out.write_text(json.dumps(asdict(summary), indent=2) + "\n")


# ---------- main --------------------------------------------------------------

def main(argv: Optional[list[str]] = None) -> int:
    p = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    src = p.add_mutually_exclusive_group()
    src.add_argument("--evaluation-json", type=Path, default=DEFAULT_EVAL_JSON,
                     help="Path to a vision_responses_evaluator JSON dump.")
    src.add_argument("--evaluation-id", type=str,
                     help="Supabase document_evaluations.id (UUID).")
    manual = p.add_mutually_exclusive_group()
    manual.add_argument("--manual-csv", type=Path, default=DEFAULT_MANUAL_CSV,
                        help="Path to the extracted manual-ratings CSV.")
    manual.add_argument("--manual-xlsx", type=Path,
                        help="Path to the source Excel workbook.")
    p.add_argument("--evaluator-column", type=str, default="Dompe (PK)",
                   help="Which manual evaluator column to compare against.")
    p.add_argument("--out-dir", type=Path, default=None,
                   help="Output directory (defaults to docs/docs_for_eval/comparison_<UTC-timestamp>).")
    args = p.parse_args(argv)

    timestamp = datetime.now(tz=timezone.utc).strftime("%Y%m%dT%H%M%SZ")
    out_dir = args.out_dir or (DEFAULT_OUT_PARENT / f"comparison_{timestamp}")
    out_dir.mkdir(parents=True, exist_ok=True)

    if args.manual_xlsx is not None:
        manual_data = load_manual_xlsx(args.manual_xlsx, args.evaluator_column)
        manual_input = str(args.manual_xlsx)
    else:
        manual_data = load_manual_csv(args.manual_csv, args.evaluator_column)
        manual_input = str(args.manual_csv)

    if args.evaluation_id:
        auto_data = load_auto_supabase(args.evaluation_id)
        auto_input = f"supabase:document_evaluations:{args.evaluation_id}"
    else:
        auto_data = load_auto_json(args.evaluation_json)
        auto_input = str(args.evaluation_json)

    rows = build_rows(manual_data, auto_data)
    if not rows:
        print("No comparable rows found; check that clauses align.", file=sys.stderr)
        return 1

    summary = summarize(rows)
    summary.generated_at = datetime.now(tz=timezone.utc).isoformat()
    summary.inputs = {
        "manual": manual_input,
        "evaluator_column": args.evaluator_column,
        "auto": auto_input,
    }

    write_markdown(rows, summary, out_dir / "comparison.md")
    write_xlsx(rows, summary, out_dir / "comparison.xlsx")
    write_summary_json(summary, out_dir / "summary.json")

    print(f"Wrote {out_dir}/")
    print(f"  total={summary.total}  exact={summary.exact} ({summary.exact_pct}%)  "
          f"adjacent={summary.adjacent} ({summary.adjacent_pct}%)  "
          f"severe={summary.severe} ({summary.severe_pct}%)  na={summary.not_applicable}")
    for r in rows:
        print(f"  {r.clause:<10} manual={r.manual_rating[:18]:<18}  "
              f"auto={r.auto_status_label:<18}  agreement={r.agreement}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
