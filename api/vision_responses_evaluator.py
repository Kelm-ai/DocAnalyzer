#!/usr/bin/env python3
"""
ISO 14971 Vision Responses Evaluator

Uploads a PDF to the selected provider's Files API (OpenAI or Gemini) once,
reuses the returned handle across parallel calls, and evaluates the ISO test
requirements. This mirrors ChatGPT's "look at the document" behaviour while
keeping output structure comparable with the markdown-based evaluator.
"""

from __future__ import annotations

import argparse
import asyncio
import hashlib
import json
import logging
import os
import tempfile
from datetime import datetime, timedelta, timezone
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

# Load environment variables from .env file
from dotenv import load_dotenv
load_dotenv()

from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from evaluation_schema import RequirementEvaluationSchema

import openai
from openai import OpenAI

try:
    from google import genai
    from google.genai import types as genai_types
    GENAI_AVAILABLE = True
except ImportError:
    GENAI_AVAILABLE = False
    genai = None  # type: ignore
    genai_types = None  # type: ignore
    print("Warning: google-genai not available. Gemini provider will be disabled.")

try:
    import anthropic
    from anthropic import Anthropic
    ANTHROPIC_AVAILABLE = True
except ImportError:
    ANTHROPIC_AVAILABLE = False
    anthropic = None  # type: ignore
    Anthropic = None  # type: ignore
    print("Warning: anthropic SDK not available. Claude provider will be disabled.")

# Import rate limiter for Claude API calls
try:
    from api.rate_limiter import get_rate_limiter
except ImportError:
    try:
        from rate_limiter import get_rate_limiter
    except ImportError:
        # Fallback: create a no-op rate limiter if module not available
        def get_rate_limiter():
            class NoOpRateLimiter:
                async def acquire(self, tokens=None):
                    return (0.0, 0)
                async def record_actual_usage(self, estimated, actual):
                    pass
                async def handle_429_error(self, retry_after=None):
                    return 10.0
            return NoOpRateLimiter()

# Retry and fallback configuration
RETRYABLE_STATUS_CODES = {429, 500, 503, 529}
MAX_RETRIES = 2
BASE_DELAY_SECONDS = 2.0
REDUCED_CONCURRENCY_FACTOR = 0.5  # Halve concurrency on first retry

# Provider-specific concurrency defaults
PROVIDER_CONCURRENCY = {
    "openai": 8,
    "claude": 4,   # Conservative due to 50 RPM limit for Opus 4.5
    "gemini": 8,
}

# Fallback mapping: if primary fails, try fallback
FALLBACK_PROVIDER = {
    "openai": "gemini",
    "claude": "gemini",
    "gemini": None,  # No fallback for Gemini
}

try:
    from supabase import create_client, Client
    SUPABASE_AVAILABLE = True
except ImportError:
    SUPABASE_AVAILABLE = False
    print("Warning: Supabase not available. Will use local requirements file.")

logger = logging.getLogger(__name__)


class VisionResponsesEvaluator:
    """Evaluate ISO requirements using gpt-5-mini with attached PDF file."""

    BASE_INSTRUCTION = """
You are an expert medical device risk-management assessor with deep knowledge of ISO 14971:2019 and ISO/TR 24971. Review ONE DOCUMENT AT A TIME and judge whether it addresses specific requirements from ISO 14971:2019 clauses 4-10.

Context and assumptions:
- Treat the document as a top-level risk-management artifact (procedure, RMP, RMR, etc.). It may reference other SOPs, work instructions, or records; clear cross-references are acceptable evidence that such systems/records exist.
- Focus on whether the document (a) defines the required process/structure and (b) shows it is practicable/implemented. Do not score clauses 1-3 or annexes as standalone requirements.

How to review each clause invocation:
1) Understand what type of document this is and how it fits the risk-management system.
2) Focus ONLY on the requested clause; search the entire document (headings, lists, tables, appendices, images/OCR) for relevant evidence of the process and expected records.
3) Presence vs adequacy: assess basic alignment with the clause. Minor ambiguity or “could be better” solutions can still PASS; treat those as opportunities for improvement (OFIs).
4) Cross-references: if the document points to another controlled SOP or record, consider that evidence that the process/record exists; do not invent details that are not written.

Decision logic (map to our schema):
- PASS: Requirement is clearly addressed; process/structure is described and evidence/records are indicated (directly or via cross-reference). Capture OFIs separately.
- FLAGGED (flag_for_review): Evidence exists but is incomplete/ambiguous or needs human confirmation; or statements conflict. Use for genuine uncertainty.
- FAIL: Core expectations are missing/contradicted; required process/records are not defined and no reasonable indication they exist.
- NOT_APPLICABLE: Use only if the clause truly does not apply to the document provided.
When in doubt between PASS and FLAGGED, prefer PASS and note OFIs; use FLAGGED only when a human needs to review.

Vision handling:
- Use both text and visual content. When graphs appear, read axis titles/units and summarise trends. When tables appear, read cells and preserve structure. If text appears in an image, transcribe it before reasoning. If something is unreadable, write "[unreadable]" and move on.
""".strip()

    def __init__(
        self,
        *,
        openai_api_key: Optional[str] = None,
        gemini_api_key: Optional[str] = None,
        model: Optional[str] = None,
        provider: Optional[str] = None,
        system_prompt: Optional[str] = None,
        framework_id: Optional[str] = None,
    ) -> None:
        # Allow custom system prompt to override class-level BASE_INSTRUCTION
        if system_prompt:
            self.BASE_INSTRUCTION = system_prompt.strip()
        self.framework_id = framework_id
        self.provider = (provider or os.getenv("VISION_PROVIDER") or "openai").strip().lower()
        if self.provider not in {"openai", "gemini", "claude"}:
            raise RuntimeError(f"Unsupported VISION_PROVIDER '{self.provider}'. Use 'openai', 'gemini', or 'claude'.")

        self.model = model
        self.openai_client: Optional[OpenAI] = None
        self.gemini_client: Optional["genai.Client"] = None
        self.claude_client: Optional["Anthropic"] = None
        self.claude_betas: List[str] = []
        self._fallback_evaluator: Optional["VisionResponsesEvaluator"] = None
        self.gemini_response_schema = None
        self.gemini_thinking_config = None
        self.gemini_media_resolution = None
        self.gemini_part_media_resolution = None
        vision_model_override = os.getenv("VISION_MODEL")

        # Evaluation controls (set early so helper methods can rely on them)
        self.concurrent_requests = int(os.getenv("VISION_EVALUATOR_CONCURRENCY", "8"))
        self.reasoning_effort = os.getenv('VISION_REASONING_EFFORT', 'medium')
        self.requirements_limit = int(os.getenv("VISION_EVALUATOR_REQUIREMENT_LIMIT", "0"))

        if self.provider == "gemini":
            if not GENAI_AVAILABLE:
                raise RuntimeError("google-genai is required for Gemini (pip install google-genai)")
            api_key = gemini_api_key or os.getenv("GEMINI_API_KEY")
            if not api_key:
                raise RuntimeError("GEMINI_API_KEY is required when VISION_PROVIDER=gemini")
            self.model = (
                self.model
                or os.getenv("GEMINI_VISION_MODEL")
                or os.getenv("GEMINI_MODEL")
                or vision_model_override
                or "gemini-3-pro-preview"
            )
            self.gemini_client = genai.Client(api_key=api_key)
            self.gemini_response_schema = self._build_gemini_schema()
            self.gemini_thinking_config = self._resolve_gemini_thinking_config()
            self.gemini_media_resolution, self.gemini_part_media_resolution = self._resolve_gemini_media_resolution()
        elif self.provider == "claude":
            if not ANTHROPIC_AVAILABLE:
                raise RuntimeError("anthropic SDK is required for Claude (pip install anthropic)")
            api_key = os.getenv("ANTHROPIC_API_KEY")
            if not api_key:
                raise RuntimeError("ANTHROPIC_API_KEY is required when VISION_PROVIDER=claude")
            self.model = (
                self.model
                or os.getenv("CLAUDE_VISION_MODEL")
                or os.getenv("CLAUDE_MODEL")
                or vision_model_override
                or "claude-opus-4-5-20251101"
            )
            self.claude_client = Anthropic(api_key=api_key)
            self.claude_betas = ["files-api-2025-04-14"]
            # Use conservative concurrency for Claude due to 50 RPM limit
            self.concurrent_requests = min(self.concurrent_requests, PROVIDER_CONCURRENCY["claude"])
        else:
            api_key = openai_api_key or os.getenv("OPENAI_API_KEY")
            if not api_key:
                raise RuntimeError("OPENAI_API_KEY is required for the vision evaluator")
            self.model = (
                self.model
                or os.getenv("OPENAI_VISION_MODEL")
                or os.getenv("OPENAI_MODEL")
                or vision_model_override
                or "gpt-5"
            )
            self.openai_client = OpenAI(api_key=api_key)

        self.supabase: Optional[Client] = None
        if SUPABASE_AVAILABLE:
            supabase_url = os.getenv("SUPABASE_URL")
            supabase_key = os.getenv("SUPABASE_SERVICE_KEY") or os.getenv("SUPABASE_SERVICE_ROLE_KEY") or os.getenv("SUPABASE_ANON_KEY")
            if supabase_url and supabase_key:
                try:
                    self.supabase = create_client(supabase_url, supabase_key)
                except Exception as exc:
                    print(f"Warning: Failed to initialize Supabase client ({exc}). Falling back to local requirements.")

        base_dir = Path(__file__).parent
        self.requirements_path = base_dir / "requirements_test.json"
        self.output_dir = base_dir / "output" / "vision_results"
        self.responses_dir = self.output_dir / "responses"
        self.output_dir.mkdir(parents=True, exist_ok=True)
        self.responses_dir.mkdir(parents=True, exist_ok=True)

        # Legacy local cache (will be phased out in favor of DB-based caching)
        self.cache_path = self.output_dir / f"uploaded_files_cache_{self.provider}.json"
        self.file_cache = self._load_cache()

        # Gemini file TTL (48 hours, with 1 hour buffer for safety)
        self.gemini_ttl_hours = 47

        if self.provider == "openai":
            logger.info(
                "VisionResponsesEvaluator initialised (provider=%s) with openai %s (%s); client=%s; model=%s; limit=%s; concurrency=%s",
                self.provider,
                getattr(openai, "__version__", "unknown"),
                getattr(openai, "__file__", "unknown"),
                type(self.openai_client).__name__ if self.openai_client else "missing",
                self.model,
                self.requirements_limit,
                self.concurrent_requests,
            )
        elif self.provider == "claude":
            logger.info(
                "VisionResponsesEvaluator initialised (provider=%s); client=%s; model=%s; limit=%s; concurrency=%s",
                self.provider,
                type(self.claude_client).__name__ if self.claude_client else "missing",
                self.model,
                self.requirements_limit,
                self.concurrent_requests,
            )
        else:
            logger.info(
                "VisionResponsesEvaluator initialised (provider=%s); client=%s; model=%s; limit=%s; concurrency=%s",
                self.provider,
                type(self.gemini_client).__name__ if self.gemini_client else "missing",
                self.model,
                self.requirements_limit,
                self.concurrent_requests,
            )


    async def evaluate_document(self, file_path: str) -> Dict:
        document_path = Path(file_path)
        if not document_path.exists():
            raise FileNotFoundError(f"File not found: {file_path}")
        if document_path.suffix.lower() != ".pdf":
            raise ValueError("Vision evaluator currently supports PDF files only")

        file_ref, file_hash, cache_hit = await self.ensure_file_ref(document_path)

        raw_requirements = self._load_requirements()
        requirements: List[Dict] = []
        for req in raw_requirements:
            req_id = str(req.get("id") or "").strip()
            title = str(req.get("title") or req.get("requirement_text") or "").strip()
            if not req_id or not title:
                logger.warning("Skipping malformed requirement row: %s", req)
                continue
            display_value = req.get("display_order")
            try:
                display_value = int(display_value)
            except (TypeError, ValueError):
                display_value = None
            fallback_order = req.get("sort_order")  # backward compatibility if present
            try:
                fallback_order = int(fallback_order)
            except (TypeError, ValueError):
                fallback_order = None
            resolved_order = display_value if display_value is not None else fallback_order if fallback_order is not None else 0
            requirements.append({
                "id": req_id,
                "clause": str(req.get("clause") or "").strip(),
                "title": title,
                "display_order": resolved_order,
                "evaluation_type": req.get("evaluation_type"),
                # Keep requirement_text for backward compatibility/context if present
                "requirement_text": req.get("requirement_text"),
            })

        requirements.sort(key=lambda r: (r.get("display_order") or 0, r.get("clause", ""), r.get("title", "")))

        if not requirements:
            raise RuntimeError("No requirements available for evaluation")

        run_id = datetime.utcnow().strftime("%Y%m%d_%H%M%S")
        run_responses_dir = self.responses_dir / run_id
        run_responses_dir.mkdir(parents=True, exist_ok=True)

        document_stats = {
            "file_name": document_path.name,
            "file_path": str(document_path.resolve()),
            "file_size_bytes": document_path.stat().st_size,
            "uploaded_file_id": file_ref.get("file_id"),
            "uploaded_file_uri": file_ref.get("file_uri"),
            "content_hash": file_hash,
            "file_id_cache_hit": cache_hit,
            "model": self.model,
            "provider": self.provider,
            "evaluated_at": datetime.utcnow().isoformat(),
            "run_id": run_id,
        }

        semaphore = asyncio.Semaphore(self.concurrent_requests)
        # Include file_hash in file_ref for retry logic
        file_ref["file_hash"] = file_hash
        tasks = [
            self._evaluate_single_requirement(
                file_ref, requirement, semaphore, run_responses_dir,
                document_path=document_path
            )
            for requirement in requirements
        ]

        evaluations = await asyncio.gather(*tasks, return_exceptions=True)

        results: List[Dict] = []
        for requirement, evaluation in zip(requirements, evaluations):
            if isinstance(evaluation, Exception):
                results.append({
                    "requirement_id": requirement["id"],
                    "status": "ERROR",
                    "confidence": "low",  # Categorical string confidence
                    "rationale": str(evaluation),
                    "evidence": [],
                    "gaps": ["Evaluation failed"],
                    "recommendations": ["Retry requirement"],
                    "tokens_used": 0,
                })
            else:
                results.append(evaluation)

        summary = self._generate_summary(document_stats, results)
        self._persist_summary(summary, run_id)
        return summary

    async def ensure_file_ref(self, document_path: Path) -> Tuple[Dict[str, str], str, bool]:
        """
        Upload the PDF once and cache the returned identifier per provider.

        Uses Supabase Storage as the permanent source of truth, with provider-specific
        file references (Gemini/OpenAI) that may expire. When a provider reference
        expires, the file is re-uploaded from Supabase Storage.
        """
        data = document_path.read_bytes()
        file_hash = hashlib.sha256(data).hexdigest()
        file_size = len(data)

        # Try DB-based caching first (if Supabase is available)
        if self.supabase is not None:
            try:
                db_result = await self._get_or_create_file_ref_from_db(
                    document_path, data, file_hash, file_size
                )
                if db_result is not None:
                    return db_result
            except Exception as exc:
                logger.warning("DB-based file caching failed, falling back to local cache: %s", exc)

        # Fallback to legacy local JSON cache
        cached_entry = self.file_cache.get(file_hash)
        if cached_entry and cached_entry.get("provider") == self.provider:
            return cached_entry, file_hash, True

        # Upload to provider
        file_meta = await self._upload_to_provider(document_path, file_hash)

        self.file_cache[file_hash] = file_meta
        self._save_cache()
        return file_meta, file_hash, False

    async def _get_or_create_file_ref_from_db(
        self,
        document_path: Path,
        data: bytes,
        file_hash: str,
        file_size: int,
    ) -> Optional[Tuple[Dict[str, str], str, bool]]:
        """
        Check DB for existing file reference. If provider ref is valid, return it.
        If expired or missing, re-upload from Supabase Storage or upload new file.
        """
        if self.supabase is None:
            return None

        # Check for existing record by file hash
        response = self.supabase.table("document_files").select("*").eq("file_hash", file_hash).execute()

        if response.data:
            record = response.data[0]

            # Check if provider reference is still valid
            if self.provider == "gemini":
                expires_at_str = record.get("gemini_expires_at")
                if expires_at_str and record.get("gemini_file_uri"):
                    expires_at = datetime.fromisoformat(expires_at_str.replace("Z", "+00:00"))
                    if expires_at > datetime.now(timezone.utc):
                        # Valid cached Gemini reference
                        logger.info("Using cached Gemini file ref for hash %s (expires %s)", file_hash[:12], expires_at_str)
                        return {
                            "provider": "gemini",
                            "file_id": record["gemini_file_id"],
                            "file_uri": record["gemini_file_uri"],
                            "file_name": record["file_name"],
                            "mime_type": record.get("mime_type") or "application/pdf",
                            "file_hash": file_hash,
                        }, file_hash, True

                # Gemini reference expired or missing - re-upload from Supabase Storage
                logger.info("Gemini file ref expired/missing for hash %s, re-uploading from storage", file_hash[:12])
                file_bytes = await self._download_from_supabase_storage(record["storage_path"])
                file_meta = await self._upload_bytes_to_gemini(file_bytes, record["file_name"], file_hash)

                # Update DB with new Gemini reference
                now = datetime.now(timezone.utc)
                self.supabase.table("document_files").update({
                    "gemini_file_id": file_meta["file_id"],
                    "gemini_file_uri": file_meta["file_uri"],
                    "gemini_uploaded_at": now.isoformat(),
                    "gemini_expires_at": (now + timedelta(hours=self.gemini_ttl_hours)).isoformat(),
                    "updated_at": now.isoformat(),
                }).eq("id", record["id"]).execute()

                file_meta["file_hash"] = file_hash
                return file_meta, file_hash, False

            elif self.provider == "openai":
                if record.get("openai_file_id"):
                    # OpenAI files don't expire the same way, but we still track them
                    logger.info("Using cached OpenAI file ref for hash %s", file_hash[:12])
                    return {
                        "provider": "openai",
                        "file_id": record["openai_file_id"],
                        "file_uri": record["openai_file_id"],
                        "file_name": record["file_name"],
                        "file_hash": file_hash,
                    }, file_hash, True

                # OpenAI reference missing - re-upload from Supabase Storage
                logger.info("OpenAI file ref missing for hash %s, re-uploading from storage", file_hash[:12])
                file_bytes = await self._download_from_supabase_storage(record["storage_path"])
                file_meta = await self._upload_bytes_to_openai(file_bytes, record["file_name"])

                # Update DB with new OpenAI reference
                now = datetime.now(timezone.utc)
                self.supabase.table("document_files").update({
                    "openai_file_id": file_meta["file_id"],
                    "openai_uploaded_at": now.isoformat(),
                    "updated_at": now.isoformat(),
                }).eq("id", record["id"]).execute()

                file_meta["file_hash"] = file_hash
                return file_meta, file_hash, False

            elif self.provider == "claude":
                if record.get("claude_file_id"):
                    # Claude files don't expire like Gemini - they persist indefinitely
                    logger.info("Using cached Claude file ref for hash %s", file_hash[:12])
                    return {
                        "provider": "claude",
                        "file_id": record["claude_file_id"],
                        "file_uri": record["claude_file_id"],
                        "file_name": record["file_name"],
                        "file_hash": file_hash,
                    }, file_hash, True

                # Claude reference missing - re-upload from Supabase Storage
                logger.info("Claude file ref missing for hash %s, re-uploading from storage", file_hash[:12])
                file_bytes = await self._download_from_supabase_storage(record["storage_path"])
                file_meta = await self._upload_bytes_to_claude(file_bytes, record["file_name"])

                # Update DB with new Claude reference
                now = datetime.now(timezone.utc)
                self.supabase.table("document_files").update({
                    "claude_file_id": file_meta["file_id"],
                    "claude_uploaded_at": now.isoformat(),
                    "updated_at": now.isoformat(),
                }).eq("id", record["id"]).execute()

                file_meta["file_hash"] = file_hash
                return file_meta, file_hash, False

        else:
            # New file - upload to Supabase Storage first, then to provider
            storage_path = await self._upload_to_supabase_storage(document_path, file_hash)
            file_meta = await self._upload_to_provider(document_path, file_hash)

            # Insert new DB record
            now = datetime.now(timezone.utc)
            db_record = {
                "storage_path": storage_path,
                "file_hash": file_hash,
                "file_name": document_path.name,
                "mime_type": "application/pdf",
                "file_size_bytes": file_size,
                "created_at": now.isoformat(),
                "updated_at": now.isoformat(),
            }

            if self.provider == "gemini":
                db_record.update({
                    "gemini_file_id": file_meta["file_id"],
                    "gemini_file_uri": file_meta["file_uri"],
                    "gemini_uploaded_at": now.isoformat(),
                    "gemini_expires_at": (now + timedelta(hours=self.gemini_ttl_hours)).isoformat(),
                })
            elif self.provider == "openai":
                db_record.update({
                    "openai_file_id": file_meta["file_id"],
                    "openai_uploaded_at": now.isoformat(),
                })
            elif self.provider == "claude":
                db_record.update({
                    "claude_file_id": file_meta["file_id"],
                    "claude_uploaded_at": now.isoformat(),
                })

            self.supabase.table("document_files").insert(db_record).execute()
            file_meta["file_hash"] = file_hash
            return file_meta, file_hash, False

        return None

    async def _upload_to_supabase_storage(self, file_path: Path, file_hash: str) -> str:
        """Upload file to Supabase Storage bucket, return storage path."""
        if self.supabase is None:
            raise RuntimeError("Supabase client is not configured")

        storage_path = f"evaluations/{file_hash}/{file_path.name}"
        data = file_path.read_bytes()

        try:
            self.supabase.storage.from_("documents").upload(
                storage_path,
                data,
                {"content-type": "application/pdf", "upsert": "true"}
            )
            logger.info("Uploaded file to Supabase Storage: %s", storage_path)
        except Exception as exc:
            # If file already exists, that's fine
            if "already exists" not in str(exc).lower() and "duplicate" not in str(exc).lower():
                raise
            logger.info("File already exists in Supabase Storage: %s", storage_path)

        return storage_path

    async def _download_from_supabase_storage(self, storage_path: str) -> bytes:
        """Download file from Supabase Storage bucket."""
        if self.supabase is None:
            raise RuntimeError("Supabase client is not configured")

        response = self.supabase.storage.from_("documents").download(storage_path)
        logger.info("Downloaded file from Supabase Storage: %s (%d bytes)", storage_path, len(response))
        return response

    async def _upload_to_provider(self, document_path: Path, file_hash: str) -> Dict[str, str]:
        """Upload file to the configured vision provider (Gemini, OpenAI, or Claude)."""
        if self.provider == "gemini":
            if self.gemini_client is None:
                raise RuntimeError("Gemini client is not configured")
            upload = await asyncio.to_thread(
                self.gemini_client.files.upload,
                file=document_path,
            )
            file_id = getattr(upload, "name", None) or getattr(upload, "uri", None)
            file_uri = getattr(upload, "uri", None) or file_id
            file_meta = {
                "provider": "gemini",
                "file_id": file_id,
                "file_uri": file_uri,
                "file_name": getattr(upload, "display_name", None) or document_path.name,
                "mime_type": getattr(upload, "mime_type", None) or "application/pdf",
                "uploaded_at": datetime.now(timezone.utc).isoformat(),
            }
            if not file_meta["file_uri"]:
                raise RuntimeError("Gemini upload failed: missing file URI")
            return file_meta
        elif self.provider == "claude":
            if self.claude_client is None:
                raise RuntimeError("Claude client is not configured")
            upload = await asyncio.to_thread(
                self.claude_client.beta.files.upload,
                file=document_path,
            )
            return {
                "provider": "claude",
                "file_id": upload.id,
                "file_uri": upload.id,
                "file_name": getattr(upload, "filename", None) or document_path.name,
                "mime_type": getattr(upload, "mime_type", None) or "application/pdf",
                "uploaded_at": datetime.now(timezone.utc).isoformat(),
            }
        else:
            if self.openai_client is None:
                raise RuntimeError("OpenAI client is not configured")
            with open(document_path, "rb") as file_obj:
                upload = await asyncio.to_thread(
                    self.openai_client.files.create,
                    file=file_obj,
                    purpose="user_data",
                )
            return {
                "provider": "openai",
                "file_id": upload.id,
                "file_uri": getattr(upload, "id", None),
                "file_name": document_path.name,
                "uploaded_at": datetime.now(timezone.utc).isoformat(),
            }

    async def _upload_bytes_to_gemini(self, file_bytes: bytes, file_name: str, file_hash: str) -> Dict[str, str]:
        """Upload bytes to Gemini (requires writing to temp file first)."""
        if self.gemini_client is None:
            raise RuntimeError("Gemini client is not configured")

        with tempfile.NamedTemporaryFile(suffix=".pdf", delete=False) as tmp:
            tmp.write(file_bytes)
            tmp_path = Path(tmp.name)

        try:
            upload = await asyncio.to_thread(
                self.gemini_client.files.upload,
                file=tmp_path,
            )
            file_id = getattr(upload, "name", None) or getattr(upload, "uri", None)
            file_uri = getattr(upload, "uri", None) or file_id
            file_meta = {
                "provider": "gemini",
                "file_id": file_id,
                "file_uri": file_uri,
                "file_name": file_name,
                "mime_type": getattr(upload, "mime_type", None) or "application/pdf",
                "uploaded_at": datetime.now(timezone.utc).isoformat(),
            }
            if not file_meta["file_uri"]:
                raise RuntimeError("Gemini upload failed: missing file URI")
            return file_meta
        finally:
            tmp_path.unlink(missing_ok=True)

    async def _upload_bytes_to_openai(self, file_bytes: bytes, file_name: str) -> Dict[str, str]:
        """Upload bytes to OpenAI (requires writing to temp file first)."""
        if self.openai_client is None:
            raise RuntimeError("OpenAI client is not configured")

        with tempfile.NamedTemporaryFile(suffix=".pdf", delete=False) as tmp:
            tmp.write(file_bytes)
            tmp_path = Path(tmp.name)

        try:
            with open(tmp_path, "rb") as file_obj:
                upload = await asyncio.to_thread(
                    self.openai_client.files.create,
                    file=file_obj,
                    purpose="user_data",
                )
            return {
                "provider": "openai",
                "file_id": upload.id,
                "file_uri": getattr(upload, "id", None),
                "file_name": file_name,
                "uploaded_at": datetime.now(timezone.utc).isoformat(),
            }
        finally:
            tmp_path.unlink(missing_ok=True)

    async def _upload_bytes_to_claude(self, file_bytes: bytes, file_name: str) -> Dict[str, str]:
        """Upload bytes to Claude Files API (requires writing to temp file first)."""
        if self.claude_client is None:
            raise RuntimeError("Claude client is not configured")

        with tempfile.NamedTemporaryFile(suffix=".pdf", delete=False) as tmp:
            tmp.write(file_bytes)
            tmp_path = Path(tmp.name)

        try:
            upload = await asyncio.to_thread(
                self.claude_client.beta.files.upload,
                file=tmp_path,
            )
            return {
                "provider": "claude",
                "file_id": upload.id,
                "file_uri": upload.id,
                "file_name": file_name,
                "mime_type": getattr(upload, "mime_type", None) or "application/pdf",
                "uploaded_at": datetime.now(timezone.utc).isoformat(),
            }
        finally:
            tmp_path.unlink(missing_ok=True)

    async def invalidate_provider_ref(self, file_hash: str) -> None:
        """
        Invalidate the provider file reference for a given file hash.
        Called when a 403 PERMISSION_DENIED error is encountered.
        """
        if self.supabase is None:
            # Clear from local cache
            if file_hash in self.file_cache:
                del self.file_cache[file_hash]
                self._save_cache()
            return

        now = datetime.now(timezone.utc)
        if self.provider == "gemini":
            self.supabase.table("document_files").update({
                "gemini_file_id": None,
                "gemini_file_uri": None,
                "gemini_expires_at": None,
                "updated_at": now.isoformat(),
            }).eq("file_hash", file_hash).execute()
        elif self.provider == "openai":
            self.supabase.table("document_files").update({
                "openai_file_id": None,
                "updated_at": now.isoformat(),
            }).eq("file_hash", file_hash).execute()
        elif self.provider == "claude":
            self.supabase.table("document_files").update({
                "claude_file_id": None,
                "claude_uploaded_at": None,
                "updated_at": now.isoformat(),
            }).eq("file_hash", file_hash).execute()

        logger.info("Invalidated %s file ref for hash %s", self.provider, file_hash[:12])

    def _error_result(self, requirement: Dict, rationale: str, gap: str, tokens_used: int = 0) -> Dict:
        """Create standardized error result."""
        return {
            "requirement_id": requirement["id"],
            "status": "ERROR",
            "confidence": "low",
            "rationale": rationale,
            "evidence": [],
            "gaps": [gap],
            "recommendations": ["Retry requirement"],
            "tokens_used": tokens_used,
        }

    def _parse_json_response(self, text: str) -> Optional[Dict]:
        """Parse JSON from response text, handling markdown code blocks."""
        if not text:
            return None
        try:
            return json.loads(text)
        except json.JSONDecodeError:
            import re
            match = re.search(r'\{[\s\S]*\}', text)
            if match:
                try:
                    return json.loads(match.group())
                except json.JSONDecodeError:
                    pass
        return None

    def _is_retryable_error(self, error_str: str) -> bool:
        """Check if an error is retryable (transient)."""
        retryable_patterns = ["429", "500", "503", "529", "overload", "rate_limit", "RateLimitError", "ServerError"]
        return any(pattern.lower() in error_str.lower() for pattern in retryable_patterns)

    def _estimate_tokens(self, text: str) -> int:
        """
        Estimate token count for text.
        Uses approximation: ~4 characters per token.
        """
        return len(text) // 4 + 100  # Add small buffer

    def _extract_retry_after(self, error_str: str) -> Optional[float]:
        """Extract retry-after value from error message if present."""
        import re
        # Look for patterns like "retry after 60 seconds" or "Retry-After: 60"
        match = re.search(r'retry.?after[:\s]+(\d+)', error_str, re.IGNORECASE)
        if match:
            return float(match.group(1))
        return None

    async def _evaluate_with_fallback(
        self,
        document_path: Path,
        requirement: Dict,
        run_responses_dir: Path,
        fallback_provider: str,
    ) -> Dict:
        """Evaluate using fallback provider (Gemini)."""
        if self._fallback_evaluator is None or self._fallback_evaluator.provider != fallback_provider:
            self._fallback_evaluator = VisionResponsesEvaluator(provider=fallback_provider)

        # Get file ref for fallback provider
        file_ref, file_hash, _ = await self._fallback_evaluator.ensure_file_ref(document_path)
        file_ref["file_hash"] = file_hash

        fallback_semaphore = asyncio.Semaphore(PROVIDER_CONCURRENCY.get(fallback_provider, 4))

        if fallback_provider == "gemini":
            result = await self._fallback_evaluator._evaluate_single_requirement_gemini(
                file_ref, requirement, fallback_semaphore, run_responses_dir
            )
        else:
            result = await self._fallback_evaluator._evaluate_single_requirement_openai(
                file_ref.get("file_id", ""), requirement, fallback_semaphore, run_responses_dir
            )

        result["fallback_provider"] = fallback_provider
        return result

    async def _evaluate_single_requirement(
        self,
        file_ref: Dict,
        requirement: Dict,
        semaphore: asyncio.Semaphore,
        run_responses_dir: Path,
        document_path: Optional[Path] = None,
        retry_count: int = 0,
    ) -> Dict:
        """Evaluate with retry and fallback logic."""
        try:
            if self.provider == "claude":
                result = await self._evaluate_single_requirement_claude(
                    file_ref, requirement, semaphore, run_responses_dir
                )
            elif self.provider == "gemini":
                result = await self._evaluate_single_requirement_gemini(
                    file_ref, requirement, semaphore, run_responses_dir
                )
            else:
                result = await self._evaluate_single_requirement_openai(
                    file_ref.get("file_id", ""), requirement, semaphore, run_responses_dir
                )

            # Check for retryable errors in result
            if result.get("status") == "ERROR":
                rationale = result.get("rationale", "")

                # Handle PERMISSION_DENIED (file expired/invalid)
                if "PERMISSION_DENIED" in rationale or "not_found" in rationale.lower():
                    if retry_count < 1 and document_path is not None:
                        logger.warning(
                            "Got file access error for requirement %s, invalidating cache and retrying...",
                            requirement["id"]
                        )
                        file_hash = file_ref.get("file_hash", "")
                        if file_hash:
                            await self.invalidate_provider_ref(file_hash)
                            new_file_ref, _, _ = await self.ensure_file_ref(document_path)
                            new_file_ref["file_hash"] = file_hash
                            return await self._evaluate_single_requirement(
                                new_file_ref, requirement, semaphore, run_responses_dir,
                                document_path=document_path, retry_count=retry_count + 1
                            )

                # Handle retryable errors (rate limits, overload, server errors)
                if self._is_retryable_error(rationale):
                    if retry_count < 1:
                        # For rate limit errors, use longer delays from rate limiter
                        if "429" in rationale or "rate_limit" in rationale.lower():
                            rate_limiter = get_rate_limiter(model=self.model)
                            delay = await rate_limiter.handle_429_error(self._extract_retry_after(rationale))
                        else:
                            delay = BASE_DELAY_SECONDS * (2 ** retry_count)  # Exponential backoff

                        logger.warning(
                            "Retryable error for requirement %s (attempt %d), retrying after %.1fs...",
                            requirement["id"], retry_count + 1, delay
                        )
                        await asyncio.sleep(delay)
                        return await self._evaluate_single_requirement(
                            file_ref, requirement, semaphore, run_responses_dir,
                            document_path=document_path, retry_count=retry_count + 1
                        )

                    # After retries exhausted, try fallback provider
                    fallback_provider = FALLBACK_PROVIDER.get(self.provider)
                    if fallback_provider and document_path is not None:
                        logger.warning(
                            "Falling back to %s for requirement %s after %s failures",
                            fallback_provider, requirement["id"], self.provider
                        )
                        return await self._evaluate_with_fallback(
                            document_path, requirement, run_responses_dir, fallback_provider
                        )

            return result

        except Exception as exc:
            error_str = str(exc)
            logger.exception("Exception during evaluation of requirement %s", requirement["id"])

            if self._is_retryable_error(error_str):
                if retry_count < 1:
                    # For rate limit errors, use longer delays from rate limiter
                    if "429" in error_str or "rate_limit" in error_str.lower():
                        rate_limiter = get_rate_limiter(model=self.model)
                        delay = await rate_limiter.handle_429_error(self._extract_retry_after(error_str))
                    else:
                        delay = BASE_DELAY_SECONDS * (2 ** retry_count)  # Exponential backoff

                    logger.warning(
                        "Retrying requirement %s after exception (attempt %d), waiting %.1fs",
                        requirement["id"], retry_count + 1, delay
                    )
                    await asyncio.sleep(delay)
                    return await self._evaluate_single_requirement(
                        file_ref, requirement, semaphore, run_responses_dir,
                        document_path=document_path, retry_count=retry_count + 1
                    )

                # Try fallback provider
                fallback_provider = FALLBACK_PROVIDER.get(self.provider)
                if fallback_provider and document_path is not None:
                    logger.warning(
                        "Falling back to %s for requirement %s after exception",
                        fallback_provider, requirement["id"]
                    )
                    return await self._evaluate_with_fallback(
                        document_path, requirement, run_responses_dir, fallback_provider
                    )

            return self._error_result(requirement, error_str, f"{self.provider} evaluation failed")

    async def _evaluate_single_requirement_openai(
        self,
        file_id: str,
        requirement: Dict,
        semaphore: asyncio.Semaphore,
        run_responses_dir: Path,
    ) -> Dict:
        async with semaphore:
            prompt = self._build_prompt(requirement)

            try:
                response = await asyncio.to_thread(
                    self.openai_client.responses.parse,  # type: ignore[union-attr]
                    model=self.model,
                    reasoning={"effort": self.reasoning_effort},
                    input=[
                        {
                            "role": "user",
                            "content": [
                                {"type": "input_text", "text": prompt},
                                {"type": "input_file", "file_id": file_id},
                            ],
                        }
                    ],
                    text_format=RequirementEvaluationSchema,
                )
            except AttributeError as attr_err:
                logger.exception("Vision API attribute error for requirement %s", requirement['id'])
                return {
                    "requirement_id": requirement["id"],
                    "status": "ERROR",
                    "confidence": "low",  # Categorical string confidence
                    "rationale": str(attr_err),
                    "evidence": [],
                    "gaps": ["OpenAI client lacks responses API"],
                    "recommendations": ["Upgrade openai package"],
                    "tokens_used": 0,
                }

            usage = getattr(response, "usage", None)
            tokens_used = getattr(usage, "total_tokens", 0) if usage else 0

            parsed_model = getattr(response, "output_parsed", None)
            if parsed_model is None:
                return {
                    "requirement_id": requirement["id"],
                    "status": "ERROR",
                    "confidence": "low",  # Categorical string confidence
                    "rationale": "Structured output missing from model response",
                    "evidence": [],
                    "gaps": ["Model response missing structured payload"],
                    "recommendations": ["Retry evaluation"],
                    "tokens_used": tokens_used,
                }

            parsed = parsed_model.model_dump()
            parsed.setdefault("requirement_id", requirement["id"])
            parsed.setdefault("requirement_title", requirement.get("title"))
            parsed.setdefault("requirement_clause", requirement.get("clause"))
            parsed["tokens_used"] = tokens_used
            raw_file = run_responses_dir / f"response_{requirement['id'].replace('-', '_')}.txt"
            raw_text = getattr(response, "output_text", None) or json.dumps(parsed, indent=2)
            raw_file.write_text(raw_text, encoding="utf-8")
            return parsed

    async def _evaluate_single_requirement_claude(
        self,
        file_ref: Dict,
        requirement: Dict,
        semaphore: asyncio.Semaphore,
        run_responses_dir: Path,
    ) -> Dict:
        """Evaluate a single requirement using Claude's vision API with rate limiting."""
        async with semaphore:
            prompt = self._build_prompt(requirement)
            file_id = file_ref.get("file_id")

            if not file_id:
                return self._error_result(requirement, "Missing file_id for Claude evaluation", "File reference missing")

            # Estimate tokens for this request (prompt + file reference overhead)
            estimated_tokens = self._estimate_tokens(prompt) + 4000  # Base overhead for PDF

            # Acquire rate limit permission (model-specific)
            rate_limiter = get_rate_limiter(model=self.model)
            wait_time, current_usage = await rate_limiter.acquire(estimated_tokens)

            if wait_time > 0:
                logger.info(
                    "Rate limiter delayed request for requirement %s by %.1fs",
                    requirement["id"], wait_time
                )

            try:
                response = await asyncio.to_thread(
                    self.claude_client.beta.messages.create,  # type: ignore[union-attr]
                    model=self.model,
                    max_tokens=4096,
                    betas=self.claude_betas,
                    messages=[{
                        "role": "user",
                        "content": [
                            {
                                "type": "document",
                                "source": {"type": "file", "file_id": file_id}
                            },
                            {"type": "text", "text": prompt},
                        ]
                    }]
                )
            except Exception as exc:
                error_str = str(exc)
                logger.exception("Claude API error for requirement %s", requirement['id'])

                # Handle 429 rate limit errors specially
                if "429" in error_str or "rate_limit" in error_str.lower():
                    retry_after = self._extract_retry_after(error_str)
                    wait_time = await rate_limiter.handle_429_error(retry_after)
                    logger.warning(
                        "Rate limit hit for requirement %s, recommended wait: %.1fs",
                        requirement["id"], wait_time
                    )

                # Re-raise retryable errors so the retry logic can handle them
                if self._is_retryable_error(error_str):
                    raise
                return self._error_result(requirement, error_str, "Claude API error")

            # Extract token usage
            usage = getattr(response, "usage", None)
            tokens_used = 0
            if usage:
                tokens_used = getattr(usage, "input_tokens", 0) + getattr(usage, "output_tokens", 0)

            # Update rate limiter with actual usage
            await rate_limiter.record_actual_usage(estimated_tokens, tokens_used)

            # Parse response
            response_text = ""
            if response.content and len(response.content) > 0:
                response_text = getattr(response.content[0], "text", "")

            parsed = self._parse_json_response(response_text)
            if parsed is None:
                raw_file = run_responses_dir / f"response_{requirement['id'].replace('-', '_')}.txt"
                raw_file.write_text(response_text or "", encoding="utf-8")
                return self._error_result(requirement, "Structured output missing from Claude response", "Parse JSON response failed", tokens_used)

            parsed.setdefault("requirement_id", requirement["id"])
            parsed.setdefault("requirement_title", requirement.get("title"))
            parsed.setdefault("requirement_clause", requirement.get("clause"))
            parsed["tokens_used"] = tokens_used

            raw_file = run_responses_dir / f"response_{requirement['id'].replace('-', '_')}.txt"
            raw_file.write_text(response_text or json.dumps(parsed, indent=2), encoding="utf-8")
            return parsed

    async def _evaluate_single_requirement_gemini(
        self,
        file_ref: Dict,
        requirement: Dict,
        semaphore: asyncio.Semaphore,
        run_responses_dir: Path,
    ) -> Dict:
        async with semaphore:
            prompt = self._build_prompt(requirement)
            file_uri = file_ref.get("file_uri") or file_ref.get("file_id")
            mime_type = file_ref.get("mime_type") or "application/pdf"

            try:
                file_part = genai_types.Part(file_data=genai_types.FileData(file_uri=file_uri, mime_type=mime_type))  # type: ignore[arg-type]
                gen_config_kwargs = {
                    "response_mime_type": "application/json",
                    "response_schema": self.gemini_response_schema,
                }
                # media_resolution at the part-level is not supported on v1beta; keep config-level only when available
                if self.gemini_media_resolution is not None:
                    gen_config_kwargs["media_resolution"] = self.gemini_media_resolution
                if self.gemini_thinking_config is not None:
                    gen_config_kwargs["thinking_config"] = self.gemini_thinking_config

                response = await asyncio.to_thread(
                    self.gemini_client.models.generate_content,  # type: ignore[union-attr]
                    model=self.model,
                    contents=[file_part, prompt],
                    config=genai_types.GenerateContentConfig(**gen_config_kwargs),
                )
            except Exception as exc:
                logger.exception("Gemini API error for requirement %s", requirement['id'])
                return {
                    "requirement_id": requirement["id"],
                    "status": "ERROR",
                    "confidence": "low",
                    "rationale": str(exc),
                    "evidence": [],
                    "gaps": ["Gemini generate_content failed"],
                    "recommendations": ["Retry requirement"],
                    "tokens_used": 0,
                }

            usage = getattr(response, "usage_metadata", None)
            tokens_used = getattr(usage, "total_token_count", 0) if usage else 0

            parsed_model = getattr(response, "parsed", None)
            parsed: Optional[Dict] = None
            if parsed_model is not None:
                parsed = parsed_model if isinstance(parsed_model, dict) else parsed_model.model_dump()
            else:
                response_text = getattr(response, "text", None)
                if response_text:
                    try:
                        parsed = json.loads(response_text)
                    except json.JSONDecodeError:
                        parsed = None

            if parsed is None:
                raw_file = run_responses_dir / f"response_{requirement['id'].replace('-', '_')}.txt"
                raw_file.write_text(getattr(response, "text", "") or "", encoding="utf-8")
                return {
                    "requirement_id": requirement["id"],
                    "status": "ERROR",
                    "confidence": "low",
                    "rationale": "Structured output missing from Gemini response",
                    "evidence": [],
                    "gaps": ["Model response missing structured payload"],
                    "recommendations": ["Retry evaluation"],
                    "tokens_used": tokens_used,
                }

            parsed.setdefault("requirement_id", requirement["id"])
            parsed.setdefault("requirement_title", requirement.get("title"))
            parsed.setdefault("requirement_clause", requirement.get("clause"))
            parsed["tokens_used"] = tokens_used

            raw_file = run_responses_dir / f"response_{requirement['id'].replace('-', '_')}.txt"
            raw_text = getattr(response, "text", None) or json.dumps(parsed, indent=2)
            raw_file.write_text(raw_text, encoding="utf-8")
            return parsed

    def _build_prompt(self, requirement: Dict) -> str:
        description = requirement.get("title") or requirement.get("requirement_text") or ""
        requirement_details = "\n".join(
            [
                f"- ID: {requirement.get('id')}",
                f"- Clause: {requirement.get('clause', '')}",
                f"- Title: {description}",
                f"- Order: {requirement.get('display_order') if requirement.get('display_order') is not None else requirement.get('sort_order')}" if requirement.get("display_order") is not None or requirement.get("sort_order") is not None else "",
                f"- Evaluation Type: {requirement.get('evaluation_type')}" if requirement.get("evaluation_type") else "",
                f"- Additional Context: {requirement.get('requirement_text')}" if requirement.get("requirement_text") and requirement.get("requirement_text") != description else "",
            ]
        ).strip()

        instruction_block = (
            "MANDATORY METHOD:\n"
            "1. Review the attached PDF for visuals (tables, charts, signatures) whenever the text layer is insufficient.\n"
            "2. Evaluate ONLY the requested clause; cite page or section references for evidence. Treat clear cross-references to other SOPs/records as evidence that those processes/records exist.\n"
            "3. Decision logic: PASS if the requirement is clearly addressed and practicable; FAIL if the process/records are missing/contradicted; FLAGGED when evidence is incomplete or genuinely uncertain; NOT_APPLICABLE only when the clause truly does not apply.\n"
            "4. Before finalising, confirm the chosen status best matches the evidence; do not default to FLAGGED when PASS or FAIL is supported.\n"
            "Respond strictly with JSON using this schema:\n"
            "{\n"
            "  \"status\": \"PASS|FAIL|FLAGGED|NOT_APPLICABLE\",\n"
            "  \"confidence\": \"low|medium|high\",\n"
            "  \"rationale\": \"Brief 1-2 sentence explanation of the decision with key citations\",\n"
            "  \"evidence\": [\"Page/Section citation with brief quote\", ...],\n"
            "  \"gaps\": [string],\n"
            "  \"recommendations\": [string]\n"
            "}\n"
            "IMPORTANT - Field definitions:\n"
            "- 'gaps': Findings/deficiencies identified in the document.\n"
            "  - For FAIL/FLAGGED: Critical gaps that caused the failure (must be addressed).\n"
            "  - For PASS: Minor opportunities for improvement (OFI) - optional enhancements. Leave empty [] if fully satisfied.\n"
            "- 'recommendations': Actionable suggestions on HOW to address the gaps/OFIs. Always provide if gaps exist.\n"
            "Keep responses concise. Avoid lengthy explanations.\n"
            "Confidence level guidelines:\n"
            "- Use \"high\" when evidence is explicit, comprehensive, and directly addresses all criteria\n"
            "- Use \"medium\" when evidence is present but incomplete, requires some inference, or has minor gaps\n"
            "- Use \"low\" when evidence is sparse, ambiguous, uncertain, or requires significant assumptions\n"
        )

        prompt_sections = [
            self.BASE_INSTRUCTION,
            "".join(instruction_block),
            "Requirement:\n" + requirement_details,
        ]

        return "\n\n".join(prompt_sections)

    def _build_gemini_schema(self):
        if not GENAI_AVAILABLE or genai_types is None:
            return None
        return genai_types.Schema(
            type=genai_types.Type.OBJECT,
            properties={
                "status": genai_types.Schema(
                    type=genai_types.Type.STRING,
                    enum=["PASS", "FAIL", "FLAGGED", "NOT_APPLICABLE"],
                ),
                "confidence": genai_types.Schema(
                    type=genai_types.Type.STRING,
                    enum=["low", "medium", "high"],
                ),
                "rationale": genai_types.Schema(type=genai_types.Type.STRING),
                "evidence": genai_types.Schema(
                    type=genai_types.Type.ARRAY,
                    items=genai_types.Schema(type=genai_types.Type.STRING),
                ),
                "gaps": genai_types.Schema(
                    type=genai_types.Type.ARRAY,
                    items=genai_types.Schema(type=genai_types.Type.STRING),
                ),
                "recommendations": genai_types.Schema(
                    type=genai_types.Type.ARRAY,
                    items=genai_types.Schema(type=genai_types.Type.STRING),
                ),
            },
            required=["status", "confidence", "rationale", "evidence", "gaps", "recommendations"],
        )
    def _extract_response_text(self, response) -> str:
        try:
            output = getattr(response, "output", None)
            if not output:
                return ""

            parts: List[str] = []
            for item in output:
                content = getattr(item, "content", None)
                if not content:
                    continue
                for chunk in content:
                    chunk_type = getattr(chunk, "type", None)
                    if chunk_type in {"output_text", "text"}:
                        parts.append(getattr(chunk, "text", ""))
            return "\n".join(filter(None, parts))
        except AttributeError:
            return ""

    def _load_requirements(self) -> List[Dict]:
        """Fetch ISO requirements from Supabase when available, otherwise use local copy."""
        if self.supabase is not None:
            try:
                query = self.supabase.table('iso_requirements').select('*')
                # Filter by framework_id if provided
                if self.framework_id:
                    query = query.eq('framework_id', self.framework_id)
                query = query.order('display_order').order('id')
                if self.requirements_limit > 0:
                    query = query.limit(self.requirements_limit)
                response = query.execute()
                if response.data:
                    return response.data
                print(f"Warning: Supabase returned no requirements for framework {self.framework_id}. Falling back to local file.")
            except Exception as exc:
                print(f"Warning: Failed to load requirements from Supabase ({exc}). Falling back to local file.")

        requirements = json.loads(self.requirements_path.read_text())
        if self.requirements_limit > 0:
            return requirements[: self.requirements_limit]
        return requirements

    def _generate_summary(self, document_stats: Dict, results: List[Dict]) -> Dict:
        status_counts: Dict[str, int] = {"PASS": 0, "FAIL": 0, "FLAGGED": 0, "NOT_APPLICABLE": 0, "ERROR": 0}
        total_tokens = 0
        for result in results:
            status = result.get("status", "ERROR")
            status_counts[status] = status_counts.get(status, 0) + 1
            total_tokens += result.get("tokens_used", 0)

        scored = len(results) - status_counts.get("ERROR", 0)
        compliance_score = (status_counts.get("PASS", 0) / scored * 100) if scored else 0

        summary = {
            "document_info": document_stats,
            "evaluation_summary": {
                "total_requirements": len(results),
                "compliance_score": round(compliance_score, 1),
                "status_counts": status_counts,
                "total_tokens_used": total_tokens,
                "estimated_cost_usd": round((total_tokens / 1_000_000) * 5, 4),
            },
            "requirements_results": results,
            "generated_at": datetime.utcnow().isoformat(),
        }
        return summary

    def _persist_summary(self, summary: Dict, run_id: str) -> None:
        json_path = self.output_dir / f"vision_evaluation_{run_id}.json"
        json_path.write_text(json.dumps(summary, indent=2), encoding="utf-8")

        excel_path = self.output_dir / f"vision_evaluation_{run_id}.xlsx"
        self._export_to_excel(summary, excel_path)

    def _export_to_excel(self, summary: Dict, excel_path: Path) -> None:
        workbook = Workbook()
        summary_sheet = workbook.active
        summary_sheet.title = "Summary"

        summary_sheet.append(["Field", "Value"])
        for key, value in summary.get("document_info", {}).items():
            summary_sheet.append([key.replace('_', ' ').title(), value])

        summary_sheet.append([])
        summary_sheet.append(["Metric", "Value"])
        evaluation_summary = summary.get("evaluation_summary", {})
        for key, value in evaluation_summary.items():
            if key == "status_counts":
                continue
            summary_sheet.append([key.replace('_', ' ').title(), value])

        summary_sheet.append([])
        summary_sheet.append(["Status", "Count"])
        for status, count in evaluation_summary.get("status_counts", {}).items():
            summary_sheet.append([status, count])

        self._auto_size_columns(summary_sheet)

        requirements_sheet = workbook.create_sheet(title="Requirements")
        headers = [
            "Requirement ID",
            "Status",
            "Confidence",
            "Rationale",
            "Evidence",
            "Gaps",
            "Recommendations",
            "Tokens Used",
        ]
        requirements_sheet.append(headers)

        for record in summary.get("requirements_results", []):
            # Normalize confidence to uppercase categorical label
            confidence_raw = record.get("confidence", "low")
            confidence_str = str(confidence_raw).strip().lower()
            if confidence_str not in ("low", "medium", "high"):
                confidence_str = "low"
            confidence_label = confidence_str.upper()

            requirements_sheet.append([
                record.get("requirement_id"),
                record.get("status"),
                confidence_label,
                record.get("rationale", ""),
                "\n".join(record.get("evidence", [])),
                "\n".join(record.get("gaps", [])),
                "\n".join(record.get("recommendations", [])),
                record.get("tokens_used", 0),
            ])

        self._auto_size_columns(requirements_sheet)
        workbook.save(excel_path)

    def _auto_size_columns(self, worksheet) -> None:
        for column_cells in worksheet.columns:
            length = max(len(str(cell.value or "")) for cell in column_cells)
            worksheet.column_dimensions[get_column_letter(column_cells[0].column)].width = min(length + 2, 80)

    def _load_cache(self) -> Dict[str, Dict[str, Any]]:
        if self.cache_path.exists():
            try:
                return json.loads(self.cache_path.read_text())
            except json.JSONDecodeError:
                return {}
        return {}

    def _save_cache(self) -> None:
        self.cache_path.write_text(json.dumps(self.file_cache, indent=2), encoding="utf-8")

    def _resolve_gemini_thinking_config(self) -> Optional["genai_types.ThinkingConfig"]:
        if not GENAI_AVAILABLE or genai_types is None:
            return None
        raw = os.getenv("GEMINI_THINKING_LEVEL") or os.getenv("VISION_THINKING_LEVEL")
        normalized = (raw or "").strip().lower()
        if not normalized:
            effort = (self.reasoning_effort or "").strip().lower()
            normalized = "low" if effort == "low" else "high"
        if normalized in {"medium", "med"}:
            normalized = "high"  # Gemini 3 treats medium as high/dynamic
        if normalized not in {"low", "high"}:
            return None
        return genai_types.ThinkingConfig(thinking_level=normalized)

    def _resolve_gemini_media_resolution(self) -> Tuple[Optional[Any], Optional[Any]]:
        """Return (global_media_resolution_enum, part_media_resolution_config)."""
        if not GENAI_AVAILABLE or genai_types is None:
            return None, None
        raw = os.getenv("GEMINI_MEDIA_RESOLUTION") or os.getenv("VISION_MEDIA_RESOLUTION") or "media_resolution_medium"
        normalized = raw.strip().lower()
        mapping = {
            "low": "MEDIA_RESOLUTION_LOW",
            "media_resolution_low": "MEDIA_RESOLUTION_LOW",
            "medium": "MEDIA_RESOLUTION_MEDIUM",
            "media_resolution_medium": "MEDIA_RESOLUTION_MEDIUM",
            "high": "MEDIA_RESOLUTION_HIGH",
            "media_resolution_high": "MEDIA_RESOLUTION_HIGH",
        }
        enum_key = mapping.get(normalized)
        if not enum_key:
            return None, None
        try:
            config_value = genai_types.MediaResolution[enum_key]
        except Exception:
            config_value = None
        try:
            part_value = genai_types.PartMediaResolution(
                level=normalized if normalized.startswith("media_resolution_") else f"media_resolution_{normalized}"
            )
        except Exception:
            part_value = None
        return config_value, part_value


class DualVisionComparator:
    """
    Run both Claude and OpenAI providers per requirement in parallel, combine conservatively:
    - Primary providers: Claude + OpenAI (run in parallel)
    - Fallback: Each provider independently falls back to Gemini on transient errors
    - Status priority: FAIL > FLAGGED/PARTIAL/ERROR > PASS > NOT_APPLICABLE
    - Confidence: take the lower (less confident) level
    - Content: on agreement use Claude content; on conflict use the content from the less-permissive status
    - Agreement flag per requirement for UI ("agreement" | "conflict")

    Flow:
    Request → Claude AND OpenAI (parallel)
    ↓ (Error 429/503/529)
    Retry after 2s delay
    ↓ (Error again)
    Fallback to Gemini FOR THE MODEL THAT PRODUCED THE ERROR
    ↓ (If both Claude and OpenAI error out even with fallback)
    Return error
    """

    STATUS_PRIORITY = {
        "FAIL": 4,
        "PARTIAL": 3,
        "FLAGGED": 3,
        "ERROR": 3,
        "PASS": 2,
        "NOT_APPLICABLE": 1,
    }
    CONFIDENCE_PRIORITY = {"low": 0, "medium": 1, "high": 2}

    def __init__(
        self,
        *,
        system_prompt: Optional[str] = None,
        framework_id: Optional[str] = None,
    ) -> None:
        self.provider = "dual"
        self.system_prompt = system_prompt
        self.framework_id = framework_id
        self.primary = VisionResponsesEvaluator(
            provider="claude",
            system_prompt=system_prompt,
            framework_id=framework_id,
        )
        self.secondary = VisionResponsesEvaluator(
            provider="openai",
            system_prompt=system_prompt,
            framework_id=framework_id,
        )
        self.model = f"{self.primary.model}+{self.secondary.model}"
        self.supabase = self.primary.supabase or self.secondary.supabase
        # Shared Gemini fallback evaluator (lazily initialized)
        self._gemini_fallback: Optional[VisionResponsesEvaluator] = None

    def _get_gemini_fallback(self) -> "VisionResponsesEvaluator":
        """Get or create a Gemini evaluator for fallback."""
        if self._gemini_fallback is None:
            self._gemini_fallback = VisionResponsesEvaluator(
                provider="gemini",
                system_prompt=self.system_prompt,
                framework_id=self.framework_id,
            )
        return self._gemini_fallback

    async def evaluate_document(self, file_path: str) -> Dict[str, Any]:
        """Evaluate document using Claude and OpenAI in parallel, with Gemini fallback."""
        # Run both providers in parallel
        claude_task = asyncio.create_task(
            self._evaluate_with_fallback(self.primary, file_path, "claude")
        )
        openai_task = asyncio.create_task(
            self._evaluate_with_fallback(self.secondary, file_path, "openai")
        )

        claude_summary, openai_summary = await asyncio.gather(
            claude_task, openai_task, return_exceptions=True
        )

        # Handle exceptions from gather
        claude_failed = isinstance(claude_summary, Exception)
        openai_failed = isinstance(openai_summary, Exception)

        if claude_failed:
            logger.error("Claude evaluation failed completely: %s", claude_summary)
            claude_summary = self._empty_summary("claude", str(claude_summary))
        if openai_failed:
            logger.error("OpenAI evaluation failed completely: %s", openai_summary)
            openai_summary = self._empty_summary("openai", str(openai_summary))

        # Check if both providers produced only errors
        claude_results = claude_summary.get("requirements_results", [])
        openai_results = openai_summary.get("requirements_results", [])

        claude_all_errors = all(r.get("status") == "ERROR" for r in claude_results) if claude_results else True
        openai_all_errors = all(r.get("status") == "ERROR" for r in openai_results) if openai_results else True

        if claude_all_errors and openai_all_errors:
            logger.error("Both Claude and OpenAI failed for all requirements")

        combined_results, agreement_map, total_tokens = self._combine_results(
            claude_summary.get("requirements_results", []),
            openai_summary.get("requirements_results", []),
        )

        status_counts: Dict[str, int] = {"PASS": 0, "FAIL": 0, "FLAGGED": 0, "NOT_APPLICABLE": 0, "ERROR": 0}
        for record in combined_results:
            status = record.get("status", "ERROR")
            status_counts[status] = status_counts.get(status, 0) + 1

        scored = len(combined_results) - status_counts.get("ERROR", 0)
        compliance_score = (status_counts.get("PASS", 0) / scored * 100) if scored else 0

        run_id = datetime.utcnow().strftime("%Y%m%d_%H%M%S_dual")
        document_info = dict(claude_summary.get("document_info", {}))
        document_info.update({
            "provider": "dual",
            "model": self.model,
            "evaluated_at": datetime.utcnow().isoformat(),
            "run_id": run_id,
            "providers_used": {
                "claude": getattr(self.primary, "model", None),
                "openai": getattr(self.secondary, "model", None),
            },
        })

        # Track which requirements used fallback
        fallback_used = {
            "claude": claude_summary.get("_fallback_used", False),
            "openai": openai_summary.get("_fallback_used", False),
        }

        summary = {
            "document_info": document_info,
            "evaluation_summary": {
                "total_requirements": len(combined_results),
                "compliance_score": round(compliance_score, 1),
                "status_counts": status_counts,
                "total_tokens_used": total_tokens,
                "estimated_cost_usd": round((total_tokens / 1_000_000) * 5, 4),
            },
            "requirements_results": combined_results,
            "agreement_by_requirement": agreement_map,
            "raw_provider_results": {
                "claude": claude_summary,
                "openai": openai_summary,
            },
            "generated_at": datetime.utcnow().isoformat(),
            "providers_used": {
                "claude": getattr(self.primary, "model", None),
                "openai": getattr(self.secondary, "model", None),
            },
            "fallback_used": fallback_used,
        }

        # Persist combined summary alongside individual runs
        try:
            self.primary._persist_summary(summary, run_id)  # type: ignore[attr-defined]
        except Exception as exc:
            logger.warning("Failed to persist dual summary: %s", exc)

        return summary

    async def _evaluate_with_fallback(
        self,
        evaluator: "VisionResponsesEvaluator",
        file_path: str,
        provider_name: str,
    ) -> Dict[str, Any]:
        """
        Evaluate document with a provider, falling back to Gemini on transient errors.

        The evaluator's own retry logic will handle:
        1. First transient error → retry after delay
        2. Second transient error → fallback to Gemini (per-requirement)

        This wrapper catches complete evaluation failures and returns error summary.
        """
        try:
            summary = await evaluator.evaluate_document(file_path)

            # Check if fallback was used for any requirement
            results = summary.get("requirements_results", [])
            fallback_used = any(r.get("fallback_provider") == "gemini" for r in results)
            summary["_fallback_used"] = fallback_used

            return summary
        except Exception as exc:
            logger.exception("Complete failure for %s provider", provider_name)
            error_summary = self._empty_summary(provider_name, str(exc))
            error_summary["_fallback_used"] = False
            return error_summary

    def _empty_summary(self, provider: str, error_msg: str) -> Dict[str, Any]:
        """Create an empty summary for when a provider completely fails."""
        return {
            "document_info": {"provider": provider, "error": error_msg},
            "evaluation_summary": {
                "total_requirements": 0,
                "compliance_score": 0,
                "status_counts": {"ERROR": 0},
                "total_tokens_used": 0,
            },
            "requirements_results": [],
            "_provider_error": error_msg,
        }

    def _combine_results(
        self,
        claude_results: List[Dict[str, Any]],
        openai_results: List[Dict[str, Any]],
    ) -> Tuple[List[Dict[str, Any]], Dict[str, str], int]:
        """Combine results from Claude and OpenAI, taking the more conservative assessment.

        When only one provider succeeds (returns non-ERROR), use that result and mark
        as 'single_provider'. This ensures we still get results even if one provider fails.
        """
        openai_by_id = {str(r.get("requirement_id")): r for r in openai_results}
        claude_by_id = {str(r.get("requirement_id")): r for r in claude_results}
        combined: List[Dict[str, Any]] = []
        agreement_map: Dict[str, str] = {}
        total_tokens = 0
        processed_ids: set = set()

        for claude_record in claude_results:
            req_id = str(claude_record.get("requirement_id"))
            processed_ids.add(req_id)
            oai_record = openai_by_id.get(req_id)

            status_a = str(claude_record.get("status", "ERROR")).upper()
            claude_is_error = status_a == "ERROR"
            claude_fallback = claude_record.get("fallback_provider") == "gemini"

            if oai_record is None:
                # Only Claude has a result for this requirement
                if claude_is_error:
                    agreement_map[req_id] = "unknown"
                else:
                    agreement_map[req_id] = "single_provider"

                combined.append({
                    **claude_record,
                    "agreement_status": agreement_map[req_id],
                    "source_provider": "claude",
                    "secondary_provider_status": {
                        "claude": status_a,
                        "openai": None,
                    },
                    "fallback_used": {
                        "claude": claude_fallback,
                        "openai": False,
                    },
                })
                total_tokens += int(claude_record.get("tokens_used", 0) or 0)
                continue

            status_b = str(oai_record.get("status", "ERROR")).upper()
            openai_is_error = status_b == "ERROR"
            openai_fallback = oai_record.get("fallback_provider") == "gemini"

            # Determine agreement status based on whether providers succeeded
            if claude_is_error and openai_is_error:
                # Both failed - unknown
                agreement_map[req_id] = "unknown"
                chosen_record = claude_record  # Doesn't matter, both are ERROR
                chosen_status = "ERROR"
                chosen_provider = "claude"
                chosen_confidence = "low"
            elif claude_is_error and not openai_is_error:
                # Only OpenAI succeeded - single_provider
                agreement_map[req_id] = "single_provider"
                chosen_record = oai_record
                chosen_status = status_b
                chosen_provider = "openai"
                chosen_confidence = str(oai_record.get("confidence", oai_record.get("confidence_level", "low"))).lower()
            elif not claude_is_error and openai_is_error:
                # Only Claude succeeded - single_provider
                agreement_map[req_id] = "single_provider"
                chosen_record = claude_record
                chosen_status = status_a
                chosen_provider = "claude"
                chosen_confidence = str(claude_record.get("confidence", claude_record.get("confidence_level", "low"))).lower()
            else:
                # Both succeeded - check for agreement or conflict
                agreement = status_a == status_b
                agreement_map[req_id] = "agreement" if agreement else "conflict"

                chosen_status = self._more_conservative_status(status_a, status_b)
                chosen_provider = "claude" if chosen_status == status_a else "openai"

                confidence_a = str(claude_record.get("confidence", claude_record.get("confidence_level", "low"))).lower()
                confidence_b = str(oai_record.get("confidence", oai_record.get("confidence_level", "low"))).lower()
                chosen_confidence = self._lower_confidence(confidence_a, confidence_b)

                if agreement:
                    chosen_record = claude_record
                else:
                    chosen_record = claude_record if chosen_provider == "claude" else oai_record

            tokens_used = int(claude_record.get("tokens_used", 0) or 0) + int(oai_record.get("tokens_used", 0) or 0)
            total_tokens += tokens_used

            combined.append({
                "requirement_id": req_id,
                "requirement_title": chosen_record.get("requirement_title") or chosen_record.get("title"),
                "requirement_clause": chosen_record.get("requirement_clause") or chosen_record.get("clause"),
                "status": chosen_status,
                "confidence": chosen_confidence,
                "confidence_level": chosen_confidence,  # for downstream normalization
                "rationale": chosen_record.get("rationale") or chosen_record.get("evaluation_rationale") or "",
                "evidence": chosen_record.get("evidence", []),
                "gaps": chosen_record.get("gaps", []),
                "recommendations": chosen_record.get("recommendations", []),
                "tokens_used": tokens_used,
                "agreement_status": agreement_map[req_id],
                "source_provider": chosen_provider,
                "secondary_provider_status": {
                    "claude": status_a,
                    "openai": status_b,
                },
                "fallback_used": {
                    "claude": claude_fallback,
                    "openai": openai_fallback,
                },
            })

        # Add any OpenAI-only results (not in Claude results)
        for oai_record in openai_results:
            req_id = str(oai_record.get("requirement_id"))
            if req_id in processed_ids:
                continue

            status_b = str(oai_record.get("status", "ERROR")).upper()
            openai_is_error = status_b == "ERROR"
            openai_fallback = oai_record.get("fallback_provider") == "gemini"

            if openai_is_error:
                agreement_map[req_id] = "unknown"
            else:
                agreement_map[req_id] = "single_provider"

            combined.append({
                **oai_record,
                "agreement_status": agreement_map[req_id],
                "source_provider": "openai",
                "secondary_provider_status": {
                    "claude": None,
                    "openai": status_b,
                },
                "fallback_used": {
                    "claude": False,
                    "openai": openai_fallback,
                },
            })
            total_tokens += int(oai_record.get("tokens_used", 0) or 0)

        return combined, agreement_map, total_tokens

    def _more_conservative_status(self, status_a: str, status_b: str) -> str:
        priority_a = self.STATUS_PRIORITY.get(status_a, 3)
        priority_b = self.STATUS_PRIORITY.get(status_b, 3)
        if priority_a == priority_b:
            return status_a
        return status_a if priority_a > priority_b else status_b

    def _lower_confidence(self, conf_a: str, conf_b: str) -> str:
        rank_a = self.CONFIDENCE_PRIORITY.get(conf_a, 0)
        rank_b = self.CONFIDENCE_PRIORITY.get(conf_b, 0)
        return conf_a if rank_a <= rank_b else conf_b


async def _async_main(file_path: str) -> None:
    evaluator = VisionResponsesEvaluator()
    summary = await evaluator.evaluate_document(file_path)

    counts = summary["evaluation_summary"]["status_counts"]
    print("\n=== Vision Evaluation Complete ===")
    print(f"Document: {summary['document_info']['file_name']}")
    print(f"Model: {summary['document_info']['model']}")
    print(f"Score: {summary['evaluation_summary']['compliance_score']:.1f}%")
    print(f"Tokens used: {summary['evaluation_summary']['total_tokens_used']}")
    print("Status counts:")
    for status, count in counts.items():
        print(f"  {status:<15} {count}")


def main() -> None:
    parser = argparse.ArgumentParser(description="ISO 14971 vision-based evaluator")
    parser.add_argument("file_path", help="Path to PDF file to evaluate")
    args = parser.parse_args()

    asyncio.run(_async_main(args.file_path))


if __name__ == "__main__":
    main()
