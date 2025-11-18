#!/usr/bin/env python3
"""
ISO 14971 Vision Responses Evaluator

Uploads a PDF to OpenAI's Files API once, reuses the returned file_id across
parallel Responses API calls to gpt-5-mini, and evaluates the three test ISO
requirements. This mirrors ChatGPT's "look at the document" behaviour while
keeping output structure comparable with the markdown-based evaluator.
"""

from __future__ import annotations

import argparse
import asyncio
import hashlib
import json
import logging
import os
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Optional, Tuple

# Load environment variables from .env file
from dotenv import load_dotenv
load_dotenv()

import openai
from openai import OpenAI
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from evaluation_schema import RequirementEvaluationSchema

try:
    from supabase import create_client, Client
    SUPABASE_AVAILABLE = True
except ImportError:
    SUPABASE_AVAILABLE = False
    print("Warning: Supabase not available. Will use local requirements file.")


class VisionResponsesEvaluator:
    """Evaluate ISO requirements using gpt-5-mini with attached PDF file."""

    BASE_INSTRUCTION = (
        "Use both the document's text and visual content. When graphs appear, read axis "
        "titles/units and summarise trends. When tables appear, read the cells and preserve "
        "structure in your answer. If text appears inside an image, transcribe it before "
        "reasoning. If something is unreadable, write '[unreadable]' and move on."
    )

    def __init__(
        self,
        *,
        openai_api_key: Optional[str] = None,
        model: Optional[str] = None,
    ) -> None:
        api_key = openai_api_key or os.getenv("OPENAI_API_KEY")
        if not api_key:
            raise RuntimeError("OPENAI_API_KEY is required for the vision evaluator")

        self.model = model or os.getenv("OPENAI_VISION_MODEL", os.getenv("OPENAI_MODEL", "gpt-5"))
        self.client = OpenAI(api_key=api_key)

        self.concurrent_requests = int(os.getenv("VISION_EVALUATOR_CONCURRENCY", "8"))
        self.reasoning_effort = os.getenv('VISION_REASONING_EFFORT', 'medium')
        self.requirements_limit = int(os.getenv("VISION_EVALUATOR_REQUIREMENT_LIMIT", "0"))

        self.supabase: Optional[Client] = None
        if SUPABASE_AVAILABLE:
            supabase_url = os.getenv("SUPABASE_URL")
            supabase_key = os.getenv("SUPABASE_ANON_KEY") or os.getenv("SUPABASE_SERVICE_ROLE_KEY")
            if supabase_url and supabase_key:
                try:
                    self.supabase = create_client(supabase_url, supabase_key)
                except Exception as exc:
                    print(f"Warning: Failed to initialize Supabase client ({exc}). Falling back to local requirements.")

        base_dir = Path(__file__).parent
        self.requirements_path = base_dir / "requirements_test.json"
        self.output_dir = base_dir / "output" / "vision_results"
        self.responses_dir = self.output_dir / "responses"
        self.output_dir.mkdir(parents=True, exist_ok=True)
        self.responses_dir.mkdir(parents=True, exist_ok=True)

        self.cache_path = self.output_dir / "uploaded_files_cache.json"
        self.file_cache = self._load_cache()

        logger.info(
            "VisionResponsesEvaluator initialised with openai %s (%s); client=%s; model=%s; limit=%s; concurrency=%s",
            getattr(openai, "__version__", "unknown"),
            getattr(openai, "__file__", "unknown"),
            type(self.client).__name__,
            self.model,
            self.requirements_limit,
            self.concurrent_requests,
        )


    async def evaluate_document(self, file_path: str) -> Dict:
        document_path = Path(file_path)
        if not document_path.exists():
            raise FileNotFoundError(f"File not found: {file_path}")
        if document_path.suffix.lower() != ".pdf":
            raise ValueError("Vision evaluator currently supports PDF files only")

        file_id, file_hash, cache_hit = await self.ensure_file_id(document_path)

        requirements = self._load_requirements()

        run_id = datetime.utcnow().strftime("%Y%m%d_%H%M%S")
        run_responses_dir = self.responses_dir / run_id
        run_responses_dir.mkdir(parents=True, exist_ok=True)

        document_stats = {
            "file_name": document_path.name,
            "file_path": str(document_path.resolve()),
            "file_size_bytes": document_path.stat().st_size,
            "uploaded_file_id": file_id,
            "content_hash": file_hash,
            "file_id_cache_hit": cache_hit,
            "model": self.model,
            "evaluated_at": datetime.utcnow().isoformat(),
            "run_id": run_id,
        }

        semaphore = asyncio.Semaphore(self.concurrent_requests)
        tasks = [
            self._evaluate_single_requirement(file_id, requirement, semaphore, run_responses_dir)
            for requirement in requirements
        ]

        evaluations = await asyncio.gather(*tasks, return_exceptions=True)

        results: List[Dict] = []
        for requirement, evaluation in zip(requirements, evaluations):
            if isinstance(evaluation, Exception):
                results.append({
                    "requirement_id": requirement["id"],
                    "status": "ERROR",
                    "confidence": "low",
                    "rationale": str(evaluation),
                    "evidence": [],
                    "gaps": ["Evaluation failed"],
                    "recommendations": ["Retry requirement"],
                    "tokens_used": 0,
                })
            else:
                results.append(evaluation)

        summary = self._generate_summary(document_stats, results)
        self._persist_summary(summary, run_id)
        return summary

    async def ensure_file_id(self, document_path: Path) -> Tuple[str, str, bool]:
        data = document_path.read_bytes()
        file_hash = hashlib.sha256(data).hexdigest()
        cached_entry = self.file_cache.get(file_hash)
        if cached_entry:
            return cached_entry["file_id"], file_hash, True

        with open(document_path, "rb") as file_obj:
            upload = await asyncio.to_thread(
                self.client.files.create,
                file=file_obj,
                purpose="user_data",
            )

        self.file_cache[file_hash] = {
            "file_id": upload.id,
            "uploaded_at": datetime.utcnow().isoformat(),
            "file_name": document_path.name,
        }
        self._save_cache()
        return upload.id, file_hash, False

    async def _evaluate_single_requirement(
        self,
        file_id: str,
        requirement: Dict,
        semaphore: asyncio.Semaphore,
        run_responses_dir: Path,
    ) -> Dict:
        async with semaphore:
            prompt = self._build_prompt(requirement)

            try:
                response = await asyncio.to_thread(
                    self.client.responses.parse,
                    model=self.model,
                    reasoning={"effort": self.reasoning_effort},
                    input=[
                        {
                            "role": "user",
                            "content": [
                                {"type": "input_text", "text": prompt},
                                {"type": "input_file", "file_id": file_id},
                            ],
                        }
                    ],
                    text_format=RequirementEvaluationSchema,
                )
            except AttributeError as attr_err:
                logger.exception("Vision API attribute error for requirement %s", requirement['id'])
                return {
                    "requirement_id": requirement["id"],
                    "status": "ERROR",
                    "confidence": "low",
                    "rationale": str(attr_err),
                    "evidence": [],
                    "gaps": ["OpenAI client lacks responses API"],
                    "recommendations": ["Upgrade openai package"],
                    "tokens_used": 0,
                }

            usage = getattr(response, "usage", None)
            tokens_used = getattr(usage, "total_tokens", 0) if usage else 0

            parsed_model = getattr(response, "output_parsed", None)
            if parsed_model is None:
                return {
                    "requirement_id": requirement["id"],
                    "status": "ERROR",
                    "confidence": "low",
                    "rationale": "Structured output missing from model response",
                    "evidence": [],
                    "gaps": ["Model response missing structured payload"],
                    "recommendations": ["Retry evaluation"],
                    "tokens_used": tokens_used,
                }

            parsed = parsed_model.model_dump()
            parsed.setdefault("requirement_id", requirement["id"])
            parsed.setdefault("requirement_title", requirement.get("title"))
            parsed.setdefault("requirement_clause", requirement.get("clause"))
            parsed["tokens_used"] = tokens_used
            raw_file = run_responses_dir / f"response_{requirement['id'].replace('-', '_')}.txt"
            raw_text = getattr(response, "output_text", None) or json.dumps(parsed, indent=2)
            raw_file.write_text(raw_text, encoding="utf-8")
            return parsed

    def _build_prompt(self, requirement: Dict) -> str:
        requirement_details = "\n".join([
            f"- ID: {requirement['id']}",
            f"- Clause: {requirement['clause']}",
            f"- Title: {requirement['title']}",
            f"- Requirement Text: {requirement['requirement_text']}",
            f"- Acceptance Criteria: {requirement['acceptance_criteria']}",
            f"- Expected Artifacts: {requirement.get('expected_artifacts', 'Not specified')}",
        ])

        instruction_block = (
            "MANDATORY METHOD:\n"
            "1. Review the attached PDF for visuals (tables, charts, signatures) whenever the text layer is insufficient.\n"
            "2. Evaluate each acceptance criterion individually; cite page or section references for your evidence.\n"
            "3. Use PASS when all criteria are clearly satisfied with explicit evidence, FAIL when evidence is clearly missing or contradictory, and FLAGGED only when the evidence is partial or genuinely uncertain.\n"
            "4. Before finalising, confirm the chosen status best matches the evidence; avoid defaulting to FLAGGED when PASS or FAIL is well supported.\n"
            "Respond strictly with JSON using this schema:\n"
            "{\n"
            "  \"status\": \"PASS|FAIL|FLAGGED|NOT_APPLICABLE\",\n"
            "  \"confidence\": \"low|medium|high\",\n"
            "  \"rationale\": \"Explain satisfied/unsatisfied criteria with citations\",\n"
            "  \"evidence\": [\"Page/Section citation with quote\", ...],\n"
            "  \"gaps\": [string],\n"
            "  \"recommendations\": [string]\n"
            "}\n"
            "Use \"high\" confidence only when evidence is explicit and comprehensive, "
            "\"medium\" when evidence is mixed but leans toward your status, and "
            "\"low\" when evidence is sparse or uncertain.\n"
        )

        prompt_sections = [
            self.BASE_INSTRUCTION,
            "".join(instruction_block),
            "Requirement:\n" + requirement_details,
        ]

        return "\n\n".join(prompt_sections)
    def _extract_response_text(self, response) -> str:
        try:
            output = getattr(response, "output", None)
            if not output:
                return ""

            parts: List[str] = []
            for item in output:
                content = getattr(item, "content", None)
                if not content:
                    continue
                for chunk in content:
                    chunk_type = getattr(chunk, "type", None)
                    if chunk_type in {"output_text", "text"}:
                        parts.append(getattr(chunk, "text", ""))
            return "\n".join(filter(None, parts))
        except AttributeError:
            return ""

    def _load_requirements(self) -> List[Dict]:
        """Fetch ISO requirements from Supabase when available, otherwise use local copy."""
        if self.supabase is not None:
            try:
                query = self.supabase.table('iso_requirements').select('*').order('id')
                if self.requirements_limit > 0:
                    query = query.limit(self.requirements_limit)
                response = query.execute()
                if response.data:
                    return response.data
                print("Warning: Supabase returned no requirements. Falling back to local file.")
            except Exception as exc:
                print(f"Warning: Failed to load requirements from Supabase ({exc}). Falling back to local file.")

        requirements = json.loads(self.requirements_path.read_text())
        if self.requirements_limit > 0:
            return requirements[: self.requirements_limit]
        return requirements

    def _generate_summary(self, document_stats: Dict, results: List[Dict]) -> Dict:
        status_counts: Dict[str, int] = {"PASS": 0, "FAIL": 0, "FLAGGED": 0, "NOT_APPLICABLE": 0, "ERROR": 0}
        total_tokens = 0
        for result in results:
            status = result.get("status", "ERROR")
            status_counts[status] = status_counts.get(status, 0) + 1
            total_tokens += result.get("tokens_used", 0)

        scored = len(results) - status_counts.get("ERROR", 0)
        compliance_score = (status_counts.get("PASS", 0) / scored * 100) if scored else 0

        summary = {
            "document_info": document_stats,
            "evaluation_summary": {
                "total_requirements": len(results),
                "compliance_score": round(compliance_score, 1),
                "status_counts": status_counts,
                "total_tokens_used": total_tokens,
                "estimated_cost_usd": round((total_tokens / 1_000_000) * 5, 4),
            },
            "requirements_results": results,
            "generated_at": datetime.utcnow().isoformat(),
        }
        return summary

    def _persist_summary(self, summary: Dict, run_id: str) -> None:
        json_path = self.output_dir / f"vision_evaluation_{run_id}.json"
        json_path.write_text(json.dumps(summary, indent=2), encoding="utf-8")

        excel_path = self.output_dir / f"vision_evaluation_{run_id}.xlsx"
        self._export_to_excel(summary, excel_path)

    def _export_to_excel(self, summary: Dict, excel_path: Path) -> None:
        workbook = Workbook()
        summary_sheet = workbook.active
        summary_sheet.title = "Summary"

        summary_sheet.append(["Field", "Value"])
        for key, value in summary.get("document_info", {}).items():
            summary_sheet.append([key.replace('_', ' ').title(), value])

        summary_sheet.append([])
        summary_sheet.append(["Metric", "Value"])
        evaluation_summary = summary.get("evaluation_summary", {})
        for key, value in evaluation_summary.items():
            if key == "status_counts":
                continue
            summary_sheet.append([key.replace('_', ' ').title(), value])

        summary_sheet.append([])
        summary_sheet.append(["Status", "Count"])
        for status, count in evaluation_summary.get("status_counts", {}).items():
            summary_sheet.append([status, count])

        self._auto_size_columns(summary_sheet)

        requirements_sheet = workbook.create_sheet(title="Requirements")
        headers = [
            "Requirement ID",
            "Status",
            "Confidence",
            "Rationale",
            "Evidence",
            "Gaps",
            "Recommendations",
            "Tokens Used",
        ]
        requirements_sheet.append(headers)

        for record in summary.get("requirements_results", []):
            requirements_sheet.append([
                record.get("requirement_id"),
                record.get("status"),
                str(record.get("confidence", "low")).upper(),
                record.get("rationale", ""),
                "\n".join(record.get("evidence", [])),
                "\n".join(record.get("gaps", [])),
                "\n".join(record.get("recommendations", [])),
                record.get("tokens_used", 0),
            ])

        self._auto_size_columns(requirements_sheet)
        workbook.save(excel_path)

    def _auto_size_columns(self, worksheet) -> None:
        for column_cells in worksheet.columns:
            length = max(len(str(cell.value or "")) for cell in column_cells)
            worksheet.column_dimensions[get_column_letter(column_cells[0].column)].width = min(length + 2, 80)

    def _load_cache(self) -> Dict[str, Dict[str, str]]:
        if self.cache_path.exists():
            try:
                return json.loads(self.cache_path.read_text())
            except json.JSONDecodeError:
                return {}
        return {}

    def _save_cache(self) -> None:
        self.cache_path.write_text(json.dumps(self.file_cache, indent=2), encoding="utf-8")


async def _async_main(file_path: str) -> None:
    evaluator = VisionResponsesEvaluator()
    summary = await evaluator.evaluate_document(file_path)

    counts = summary["evaluation_summary"]["status_counts"]
    print("\n=== Vision Evaluation Complete ===")
    print(f"Document: {summary['document_info']['file_name']}")
    print(f"Model: {summary['document_info']['model']}")
    print(f"Score: {summary['evaluation_summary']['compliance_score']:.1f}%")
    print(f"Tokens used: {summary['evaluation_summary']['total_tokens_used']}")
    print("Status counts:")
    for status, count in counts.items():
        print(f"  {status:<15} {count}")


def main() -> None:
    parser = argparse.ArgumentParser(description="ISO 14971 vision-based evaluator")
    parser.add_argument("file_path", help="Path to PDF file to evaluate")
    args = parser.parse_args()

    asyncio.run(_async_main(args.file_path))


if __name__ == "__main__":
    main()

logger = logging.getLogger(__name__)
